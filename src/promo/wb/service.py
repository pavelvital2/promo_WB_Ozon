from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import timedelta
from decimal import Decimal, ROUND_CEILING, InvalidOperation
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Iterable
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from promo.file_storage.service import FileStorageService
from promo.runs.contracts import RunExecutionResult
from promo.shared.clock import utc_now
from promo.shared.contracts.audit import RunDetailAuditDTO
from promo.shared.contracts.runs import RunDTO, RunFileDTO
from promo.shared.contracts.stores import StoreDTO
from promo.shared.enums import BusinessResult, FileRole, Severity
from promo.shared.errors import ValidationFailedError
from promo.shared.files.policies import RunFileRetentionPolicy

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PRICE_ARTICLE_HEADER = "Артикул WB"
PRICE_CURRENT_PRICE_HEADER = "Текущая цена"
PRICE_NEW_DISCOUNT_HEADER = "Новая скидка"
PROMO_ARTICLE_HEADER = "Артикул WB"
PROMO_PLAN_PRICE_HEADER = "Плановая цена для акции"
PROMO_UPLOAD_DISCOUNT_HEADER = "Загружаемая скидка для участия в акции"
SUPPORTED_WORKBOOK_SUFFIX = ".xlsx"
MACRO_ENTRY = "xl/vbaProject.bin"
EXTERNAL_LINKS_PREFIX = "xl/externalLinks/"


@dataclass(slots=True, frozen=True)
class WBPromoAggregate:
    min_discount: int
    max_plan_price: Decimal


@dataclass(slots=True, frozen=True)
class WBDecision:
    row_number: int
    article: str | None
    current_price: Decimal | None
    min_discount: int | None
    max_plan_price: Decimal | None
    calculated_discount: int | None
    final_discount_pre_threshold: int | None
    final_discount: int | None
    severity: str
    decision_reason: str
    message: str
    should_write: bool


class WBExecutionStrategy:
    def __init__(self, file_storage: FileStorageService) -> None:
        self._file_storage = file_storage

    def validate(self, run: RunDTO, store: StoreDTO, files: tuple[RunFileDTO, ...]) -> RunExecutionResult:
        return self._process(run, store, files, write_output=False)

    def execute(self, run: RunDTO, store: StoreDTO, files: tuple[RunFileDTO, ...]) -> RunExecutionResult:
        return self._process(run, store, files, write_output=True)

    def _process(self, run: RunDTO, store: StoreDTO, files: tuple[RunFileDTO, ...], write_output: bool) -> RunExecutionResult:
        if len(files) < 2:
            raise ValidationFailedError("WB run requires one price file and at least one promo file", {"files": len(files)})

        price_files = [item for item in files if item.file_role == FileRole.WB_PRICE_INPUT.value]
        promo_files = [item for item in files if item.file_role == FileRole.WB_PROMO_INPUT.value]
        if len(price_files) != 1:
            raise ValidationFailedError("WB run requires exactly one price file", {"price_files": len(price_files)})
        if not promo_files:
            raise ValidationFailedError("WB run requires at least one promo file", {"promo_files": len(promo_files)})

        if store.wb_threshold_percent is None or store.wb_fallback_no_promo_percent is None or store.wb_fallback_over_threshold_percent is None:
            raise ValidationFailedError(
                "WB store settings are required",
                {
                    "store_id": store.id,
                },
            )

        price_context = self._load_price_workbook(price_files[0])
        promo_contexts = [self._load_promo_workbook(item) for item in promo_files]

        promo_aggregates, invalid_promo_rows, invalid_promo_files = self._aggregate_promos(promo_contexts)
        if not promo_aggregates and invalid_promo_files == len(promo_contexts):
            raise ValidationFailedError("All WB promo files are invalid", {"invalid_promo_files": invalid_promo_files})

        price_column = price_context.headers[PRICE_NEW_DISCOUNT_HEADER]
        decisions = self._build_price_decisions(price_context, promo_aggregates, store)
        summary = self._build_summary(run, decisions, invalid_promo_rows, invalid_promo_files)
        detail_rows = tuple(self._build_detail_row(run.id, decision) for decision in decisions)

        result_file = None
        if write_output:
            result_file = self._write_output_file(run, price_context, price_column, decisions)

        business_result = self._pick_business_result(run.operation_type, summary["warnings"])
        short_result_text = (
            f"WB {run.operation_type}: processed {summary['processed_rows']} rows, "
            f"found {summary['found_in_promos']} in promos, warnings {summary['warnings']}"
        )
        return RunExecutionResult(
            business_result=business_result,
            short_result_text=short_result_text,
            summary_json=summary,
            detail_rows=detail_rows,
            result_file=result_file,
        )

    def _load_price_workbook(self, file: RunFileDTO) -> "_WorkbookContext":
        return self._load_workbook(file, required_headers=(PRICE_ARTICLE_HEADER, PRICE_CURRENT_PRICE_HEADER, PRICE_NEW_DISCOUNT_HEADER))

    def _load_promo_workbook(self, file: RunFileDTO) -> "_WorkbookContext":
        return self._load_workbook(file, required_headers=(PROMO_ARTICLE_HEADER, PROMO_PLAN_PRICE_HEADER, PROMO_UPLOAD_DISCOUNT_HEADER))

    def _load_workbook(self, file: RunFileDTO, required_headers: tuple[str, ...]) -> "_WorkbookContext":
        path = self._file_storage.root_path / file.storage_relative_path
        self._validate_supported_workbook(file)
        archive_entries = self._inspect_archive(file, path)
        self._validate_unsafe_archive_entries(file, archive_entries)
        try:
            workbook = load_workbook(path, data_only=False)
        except Exception as exc:  # noqa: BLE001
            raise ValidationFailedError("WB workbook cannot be opened", {"file": file.original_filename, "error": str(exc)}) from exc
        sheet = workbook.worksheets[0]
        headers = self._headers_from_sheet(sheet)
        missing = [header for header in required_headers if header not in headers]
        if missing:
            raise ValidationFailedError(
                "WB workbook is missing required columns",
                {"file": file.original_filename, "missing_headers": missing},
            )
        return _WorkbookContext(file=file, workbook=workbook, sheet=sheet, headers=headers, archive_entries=archive_entries)

    def _headers_from_sheet(self, sheet) -> dict[str, int]:
        headers: dict[str, int] = {}
        for cell in sheet[1]:
            if cell.value is None:
                continue
            name = str(cell.value).strip()
            if name and name not in headers:
                headers[name] = cell.column
        return headers

    def _aggregate_promos(self, promo_contexts: Iterable["_WorkbookContext"]) -> tuple[dict[str, WBPromoAggregate], int, int]:
        aggregates: dict[str, WBPromoAggregate] = {}
        invalid_rows = 0
        invalid_files = 0
        for context in promo_contexts:
            file_has_valid_rows = False
            for row_number in range(2, context.sheet.max_row + 1):
                article = self._normalize_article(context.sheet.cell(row=row_number, column=context.headers[PROMO_ARTICLE_HEADER]).value)
                plan_price = self._normalize_decimal(context.sheet.cell(row=row_number, column=context.headers[PROMO_PLAN_PRICE_HEADER]).value)
                discount = self._normalize_discount(context.sheet.cell(row=row_number, column=context.headers[PROMO_UPLOAD_DISCOUNT_HEADER]).value)
                if article is None or plan_price is None or discount is None:
                    invalid_rows += 1
                    continue
                file_has_valid_rows = True
                existing = aggregates.get(article)
                if existing is None:
                    aggregates[article] = WBPromoAggregate(min_discount=discount, max_plan_price=plan_price)
                    continue
                aggregates[article] = WBPromoAggregate(
                    min_discount=min(existing.min_discount, discount),
                    max_plan_price=max(existing.max_plan_price, plan_price),
                )
            if not file_has_valid_rows:
                invalid_files += 1
        return aggregates, invalid_rows, invalid_files

    def _build_price_decisions(
        self,
        price_context: "_WorkbookContext",
        promo_aggregates: dict[str, WBPromoAggregate],
        store: StoreDTO,
    ) -> list[WBDecision]:
        decisions: list[WBDecision] = []
        seen_articles: set[str] = set()
        for row_number in range(2, price_context.sheet.max_row + 1):
            article = self._normalize_article(price_context.sheet.cell(row=row_number, column=price_context.headers[PRICE_ARTICLE_HEADER]).value)
            current_price = self._normalize_decimal(price_context.sheet.cell(row=row_number, column=price_context.headers[PRICE_CURRENT_PRICE_HEADER]).value)
            if article is not None:
                if article in seen_articles:
                    raise ValidationFailedError("Duplicate WB article in price file", {"article": article, "row_number": row_number})
                seen_articles.add(article)
            if article is None:
                decisions.append(
                    WBDecision(
                        row_number=row_number,
                        article=None,
                        current_price=current_price,
                        min_discount=None,
                        max_plan_price=None,
                        calculated_discount=None,
                        final_discount_pre_threshold=None,
                        final_discount=None,
                        severity=Severity.WARNING.value,
                        decision_reason="missing_article",
                        message="Артикул WB отсутствует",
                        should_write=False,
                    )
                )
                continue
            if current_price is None or current_price <= 0:
                decisions.append(
                    WBDecision(
                        row_number=row_number,
                        article=article,
                        current_price=current_price,
                        min_discount=None,
                        max_plan_price=None,
                        calculated_discount=None,
                        final_discount_pre_threshold=None,
                        final_discount=None,
                        severity=Severity.WARNING.value,
                        decision_reason="missing_current_price",
                        message="Текущая цена отсутствует или некорректна",
                        should_write=False,
                    )
                )
                continue

            aggregate = promo_aggregates.get(article)
            if aggregate is None:
                final_discount = self._clamp_percent(store.wb_fallback_no_promo_percent)
                decisions.append(
                    WBDecision(
                        row_number=row_number,
                        article=article,
                        current_price=current_price,
                        min_discount=None,
                        max_plan_price=None,
                        calculated_discount=None,
                        final_discount_pre_threshold=None,
                        final_discount=final_discount,
                        severity=Severity.INFO.value,
                        decision_reason="fallback_no_promo",
                        message="Артикул отсутствует в акциях, применён fallback_no_promo",
                        should_write=True,
                    )
                )
                continue

            calculated_discount = self._calculate_discount(current_price, aggregate.max_plan_price)
            final_discount_pre_threshold = min(aggregate.min_discount, calculated_discount)
            if final_discount_pre_threshold > store.wb_threshold_percent:
                final_discount = self._clamp_percent(store.wb_fallback_over_threshold_percent)
                reason = "fallback_over_threshold"
                severity = Severity.WARNING.value
                message = "Итог превысил threshold, применён fallback_over_threshold"
            else:
                final_discount = self._clamp_percent(final_discount_pre_threshold)
                reason = "min_discount" if aggregate.min_discount <= calculated_discount else "calculated_discount"
                severity = Severity.INFO.value
                message = "Итог рассчитан по WB правилам"
            decisions.append(
                WBDecision(
                    row_number=row_number,
                    article=article,
                    current_price=current_price,
                    min_discount=aggregate.min_discount,
                    max_plan_price=aggregate.max_plan_price,
                    calculated_discount=calculated_discount,
                    final_discount_pre_threshold=final_discount_pre_threshold,
                    final_discount=final_discount,
                    severity=severity,
                    decision_reason=reason,
                    message=message,
                    should_write=True,
                )
            )
        return decisions

    def _build_summary(self, run: RunDTO, decisions: list[WBDecision], invalid_promo_rows: int, invalid_promo_files: int) -> dict[str, object]:
        processed_rows = sum(1 for decision in decisions if decision.should_write)
        row_warnings = sum(1 for decision in decisions if decision.severity == Severity.WARNING.value)
        warnings = row_warnings + invalid_promo_rows + invalid_promo_files
        found_in_promos = sum(1 for decision in decisions if decision.decision_reason not in {"missing_article", "missing_current_price", "fallback_no_promo"})
        not_found_in_promos = sum(1 for decision in decisions if decision.decision_reason == "fallback_no_promo")
        fallback_no_promo = not_found_in_promos
        fallback_over_threshold = sum(1 for decision in decisions if decision.decision_reason == "fallback_over_threshold")
        final_from_min_discount = sum(1 for decision in decisions if decision.decision_reason == "min_discount")
        final_from_calculated_discount = sum(1 for decision in decisions if decision.decision_reason == "calculated_discount")
        skipped_rows = sum(1 for decision in decisions if not decision.should_write)
        return {
            "module_code": run.module_code,
            "operation_type": run.operation_type,
            "total_rows": len(decisions),
            "processed_rows": processed_rows,
            "warnings": warnings,
            "skipped_rows": skipped_rows,
            "found_in_promos": found_in_promos,
            "not_found_in_promos": not_found_in_promos,
            "fallback_no_promo": fallback_no_promo,
            "fallback_over_threshold": fallback_over_threshold,
            "final_from_min_discount": final_from_min_discount,
            "final_from_calculated_discount": final_from_calculated_discount,
            "invalid_promo_rows": invalid_promo_rows,
            "invalid_promo_files": invalid_promo_files,
        }

    def _build_detail_row(self, run_id: int, decision: WBDecision) -> RunDetailAuditDTO:
        payload = {
            "article": decision.article,
            "current_price": self._decimal_to_str(decision.current_price),
            "min_discount": decision.min_discount,
            "max_plan_price": self._decimal_to_str(decision.max_plan_price),
            "calculated_discount": decision.calculated_discount,
            "final_discount_pre_threshold": decision.final_discount_pre_threshold,
            "final_discount": decision.final_discount,
            "should_write": decision.should_write,
        }
        return RunDetailAuditDTO(
            id=0,
            run_id=run_id,
            row_number=decision.row_number,
            entity_key_1=decision.article,
            entity_key_2=None,
            severity=decision.severity,
            decision_reason=decision.decision_reason,
            message=decision.message,
            audit_payload_json=payload,
            created_at_utc=utc_now(),
        )

    def _write_output_file(
        self,
        run: RunDTO,
        price_context: "_WorkbookContext",
        price_column: int,
        decisions: list[WBDecision],
    ) -> RunFileDTO:
        self._assert_safe_write_context(price_context, price_column)
        for decision in decisions:
            if decision.should_write and decision.final_discount is not None:
                price_context.sheet.cell(row=decision.row_number, column=price_column).value = int(decision.final_discount)

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_path = Path(temp_file.name)
        try:
            try:
                price_context.workbook.save(temp_path)
            except Exception as exc:  # noqa: BLE001
                raise ValidationFailedError(
                    "WB workbook cannot be saved safely",
                    {"file": price_context.file.original_filename, "error": str(exc)},
                ) from exc
            stored = self._file_storage.copy_to_run_output(
                source_path=temp_path,
                module_code=run.module_code,
                store_id=run.store_id,
                public_run_number=run.public_run_number,
                original_filename=price_context.file.original_filename,
                mime_type=XLSX_MIME,
                created_at_utc=utc_now(),
            )
        finally:
            temp_path.unlink(missing_ok=True)
        expires_at = utc_now() + timedelta(days=RunFileRetentionPolicy().ttl_days)
        return RunFileDTO(
            id=0,
            run_id=run.id,
            file_role=FileRole.WB_RESULT_OUTPUT.value,
            original_filename=price_context.file.original_filename,
            stored_filename=stored.stored_filename,
            storage_relative_path=stored.storage_relative_path,
            mime_type=stored.mime_type,
            file_size_bytes=stored.file_size_bytes,
            file_sha256=stored.file_sha256,
            uploaded_at_utc=stored.created_at_utc,
            expires_at_utc=expires_at,
            is_available=True,
            unavailable_reason=None,
            created_at_utc=stored.created_at_utc,
        )

    def _calculate_discount(self, current_price: Decimal, max_plan_price: Decimal) -> int:
        ratio = (Decimal("1") - (max_plan_price / current_price)) * Decimal("100")
        result = ratio.to_integral_value(rounding=ROUND_CEILING)
        return self._clamp_percent(int(result))

    def _normalize_article(self, value: object) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        if text.endswith(".0"):
            text = text[:-2]
        return text or None

    def _normalize_decimal(self, value: object) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
        if not text:
            return None
        if text == "*":
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    def _normalize_discount(self, value: object) -> int | None:
        decimal_value = self._normalize_decimal(value)
        if decimal_value is None:
            return None
        try:
            return self._clamp_percent(int(decimal_value.to_integral_value(rounding=ROUND_CEILING)))
        except Exception:  # noqa: BLE001
            return None

    def _clamp_percent(self, value: int) -> int:
        return max(0, min(100, int(value)))

    def _decimal_to_str(self, value: Decimal | None) -> str | None:
        if value is None:
            return None
        normalized = value.normalize()
        return format(normalized, "f")

    def _pick_business_result(self, operation_type: str, warnings: int) -> str:
        if operation_type == "check":
            return BusinessResult.CHECK_PASSED_WITH_WARNINGS.value if warnings else BusinessResult.CHECK_PASSED.value
        return BusinessResult.COMPLETED_WITH_WARNINGS.value if warnings else BusinessResult.COMPLETED.value

    def _validate_supported_workbook(self, file: RunFileDTO) -> None:
        suffix = Path(file.original_filename).suffix.casefold()
        if suffix != SUPPORTED_WORKBOOK_SUFFIX:
            raise ValidationFailedError(
                "WB workbook format is unsupported",
                {"file": file.original_filename, "expected_suffix": SUPPORTED_WORKBOOK_SUFFIX, "actual_suffix": suffix or None},
            )

    def _inspect_archive(self, file: RunFileDTO, path: Path) -> tuple[str, ...]:
        try:
            with ZipFile(path) as archive:
                return tuple(item.filename for item in archive.infolist())
        except BadZipFile as exc:
            raise ValidationFailedError("WB workbook cannot be opened", {"file": file.original_filename, "error": str(exc)}) from exc

    def _validate_unsafe_archive_entries(self, file: RunFileDTO, archive_entries: tuple[str, ...]) -> None:
        normalized = {entry.replace("\\", "/") for entry in archive_entries}
        if MACRO_ENTRY in normalized:
            raise ValidationFailedError("WB workbook macros are not supported", {"file": file.original_filename})
        if any(entry.startswith(EXTERNAL_LINKS_PREFIX) for entry in normalized):
            raise ValidationFailedError("WB workbook contains external links", {"file": file.original_filename})

    def _assert_safe_write_context(self, context: "_WorkbookContext", price_column: int) -> None:
        workbook_security = getattr(context.workbook, "security", None)
        if workbook_security is not None and any(
            bool(getattr(workbook_security, attr, False))
            for attr in ("lockStructure", "lockWindows", "lockRevision")
        ):
            raise ValidationFailedError("WB workbook is protected and cannot be saved safely", {"file": context.file.original_filename})
        sheet_protection = getattr(context.sheet, "protection", None)
        if sheet_protection is not None and bool(getattr(sheet_protection, "sheet", False)):
            raise ValidationFailedError("WB worksheet is protected and cannot be saved safely", {"file": context.file.original_filename})
        for row_number in range(2, context.sheet.max_row + 1):
            cell = context.sheet.cell(row=row_number, column=price_column)
            if getattr(cell, "data_type", None) == "f":
                raise ValidationFailedError(
                    "WB workbook has formulas in writable column and cannot be saved safely",
                    {"file": context.file.original_filename, "row_number": row_number, "column_header": PRICE_NEW_DISCOUNT_HEADER},
                )


@dataclass(slots=True, frozen=True)
class _WorkbookContext:
    file: RunFileDTO
    workbook: object
    sheet: object
    headers: dict[str, int]
    archive_entries: tuple[str, ...]
