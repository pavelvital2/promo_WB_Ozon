from __future__ import annotations

from dataclasses import dataclass
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Iterable
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from promo.file_storage.service import FileStorageService
from promo.runs.contracts import RunExecutionResult
from promo.shared.clock import utc_now
from promo.shared.contracts.audit import RunDetailAuditDTO
from promo.shared.contracts.runs import RunDTO, RunFileDTO
from promo.shared.contracts.stores import StoreDTO
from promo.shared.enums import BusinessResult, FileRole, Severity
from promo.shared.errors import ValidationFailedError
from promo.shared.files.policies import RunFileRetentionPolicy

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
OZON_SHEET_NAME = "Товары и цены"
MIN_PRICE_COL = 10
PARTICIPATION_COL = 11
FINAL_PRICE_COL = 12
MIN_BOOST_COL = 15
MAX_BOOST_COL = 16
STOCK_COL = 18
SUPPORTED_WORKBOOK_SUFFIX = ".xlsx"
MACRO_ENTRY = "xl/vbaProject.bin"
EXTERNAL_LINKS_PREFIX = "xl/externalLinks/"


@dataclass(slots=True, frozen=True)
class OzonDecision:
    row_number: int
    entity_key_1: str | None
    entity_key_2: str | None
    min_price: Decimal | None
    min_boost: Decimal | None
    max_boost: Decimal | None
    stock: Decimal | None
    final_participation: bool
    final_promo_price: Decimal | None
    severity: str
    decision_reason: str
    message: str


class OzonExecutionStrategy:
    def __init__(self, file_storage: FileStorageService) -> None:
        self._file_storage = file_storage

    def validate(self, run: RunDTO, store: StoreDTO, files: tuple[RunFileDTO, ...]) -> RunExecutionResult:
        return self._process(run, store, files, write_output=False)

    def execute(self, run: RunDTO, store: StoreDTO, files: tuple[RunFileDTO, ...]) -> RunExecutionResult:
        return self._process(run, store, files, write_output=True)

    def _process(self, run: RunDTO, store: StoreDTO, files: tuple[RunFileDTO, ...], write_output: bool) -> RunExecutionResult:
        if len(files) != 1:
            raise ValidationFailedError("Ozon run requires exactly one file", {"files": len(files)})

        input_file = files[0]
        workbook_context = self._load_workbook(input_file)
        decisions = self._build_decisions(workbook_context)
        summary = self._build_summary(run, decisions)
        detail_rows = tuple(self._build_detail_row(run.id, decision) for decision in decisions)

        result_file = None
        if write_output:
            result_file = self._write_output_file(run, workbook_context, decisions)

        business_result = self._pick_business_result(run.operation_type, summary["warnings"])
        short_result_text = (
            f"Ozon {run.operation_type}: changed {summary['changed_rows']} rows, "
            f"participating {summary['participating']}, warnings {summary['warnings']}"
        )
        return RunExecutionResult(
            business_result=business_result,
            short_result_text=short_result_text,
            summary_json=summary,
            detail_rows=detail_rows,
            result_file=result_file,
        )

    def _load_workbook(self, file: RunFileDTO) -> "_WorkbookContext":
        path = self._file_storage.root_path / file.storage_relative_path
        self._validate_supported_workbook(file)
        archive_entries = self._inspect_archive(file, path)
        self._validate_unsafe_archive_entries(file, archive_entries)
        try:
            workbook = load_workbook(path, data_only=False)
        except Exception as exc:  # noqa: BLE001
            raise ValidationFailedError("Ozon workbook cannot be opened", {"file": file.original_filename, "error": str(exc)}) from exc
        if OZON_SHEET_NAME not in workbook.sheetnames:
            raise ValidationFailedError("Ozon workbook is missing required sheet", {"file": file.original_filename, "sheet": OZON_SHEET_NAME})
        sheet = workbook[OZON_SHEET_NAME]
        headers = self._headers_from_sheet(sheet)
        required_letters = {"J": MIN_PRICE_COL, "K": PARTICIPATION_COL, "L": FINAL_PRICE_COL, "O": MIN_BOOST_COL, "P": MAX_BOOST_COL, "R": STOCK_COL}
        missing = [letter for letter, column in required_letters.items() if sheet.cell(row=2, column=column).value is None]
        if missing:
            raise ValidationFailedError("Ozon workbook is missing required columns", {"file": file.original_filename, "missing_columns": missing})
        return _WorkbookContext(file=file, workbook=workbook, sheet=sheet, headers=headers, archive_entries=archive_entries)

    def _headers_from_sheet(self, sheet) -> dict[str, int]:
        headers: dict[str, int] = {}
        for cell in sheet[2]:
            if cell.value is None:
                continue
            name = str(cell.value).strip()
            if name and name not in headers:
                headers[name] = cell.column
        return headers

    def _build_decisions(self, context: "_WorkbookContext") -> list[OzonDecision]:
        decisions: list[OzonDecision] = []
        for row_number in range(4, context.sheet.max_row + 1):
            entity_key_1, entity_key_2 = self._entity_keys(context, row_number)
            min_price = self._normalize_decimal(context.sheet.cell(row=row_number, column=MIN_PRICE_COL).value)
            min_boost = self._normalize_decimal(context.sheet.cell(row=row_number, column=MIN_BOOST_COL).value)
            max_boost = self._normalize_decimal(context.sheet.cell(row=row_number, column=MAX_BOOST_COL).value)
            stock = self._normalize_decimal(context.sheet.cell(row=row_number, column=STOCK_COL).value)

            if min_price is None:
                decisions.append(
                    OzonDecision(
                        row_number=row_number,
                        entity_key_1=entity_key_1,
                        entity_key_2=entity_key_2,
                        min_price=None,
                        min_boost=min_boost,
                        max_boost=max_boost,
                        stock=stock,
                        final_participation=False,
                        final_promo_price=None,
                        severity=Severity.WARNING.value,
                        decision_reason="missing_min_price",
                        message="Минимальная цена отсутствует",
                    )
                )
                continue
            if stock is None or stock <= 0:
                decisions.append(
                    OzonDecision(
                        row_number=row_number,
                        entity_key_1=entity_key_1,
                        entity_key_2=entity_key_2,
                        min_price=min_price,
                        min_boost=min_boost,
                        max_boost=max_boost,
                        stock=stock,
                        final_participation=False,
                        final_promo_price=None,
                        severity=Severity.WARNING.value,
                        decision_reason="no_stock",
                        message="Остаток на складе отсутствует",
                    )
                )
                continue
            if min_boost is None and max_boost is None:
                decisions.append(
                    OzonDecision(
                        row_number=row_number,
                        entity_key_1=entity_key_1,
                        entity_key_2=entity_key_2,
                        min_price=min_price,
                        min_boost=None,
                        max_boost=None,
                        stock=stock,
                        final_participation=False,
                        final_promo_price=None,
                        severity=Severity.WARNING.value,
                        decision_reason="no_boost_prices",
                        message="Отсутствуют цены для буста",
                    )
                )
                continue
            if max_boost is not None and max_boost >= min_price:
                decisions.append(
                    OzonDecision(
                        row_number=row_number,
                        entity_key_1=entity_key_1,
                        entity_key_2=entity_key_2,
                        min_price=min_price,
                        min_boost=min_boost,
                        max_boost=max_boost,
                        stock=stock,
                        final_participation=True,
                        final_promo_price=max_boost,
                        severity=Severity.INFO.value,
                        decision_reason="use_max_boost_price",
                        message="Применена максимальная цена буста",
                    )
                )
                continue
            if max_boost is not None and min_boost is not None and max_boost < min_price and min_boost >= min_price:
                decisions.append(
                    OzonDecision(
                        row_number=row_number,
                        entity_key_1=entity_key_1,
                        entity_key_2=entity_key_2,
                        min_price=min_price,
                        min_boost=min_boost,
                        max_boost=max_boost,
                        stock=stock,
                        final_participation=True,
                        final_promo_price=min_price,
                        severity=Severity.INFO.value,
                        decision_reason="use_min_price",
                        message="Применена минимальная цена",
                    )
                )
                continue
            if min_boost is not None and min_boost < min_price:
                decisions.append(
                    OzonDecision(
                        row_number=row_number,
                        entity_key_1=entity_key_1,
                        entity_key_2=entity_key_2,
                        min_price=min_price,
                        min_boost=min_boost,
                        max_boost=max_boost,
                        stock=stock,
                        final_participation=False,
                        final_promo_price=None,
                        severity=Severity.WARNING.value,
                        decision_reason="below_min_price_threshold",
                        message="Цена буста ниже минимального порога",
                    )
                )
                continue
            decisions.append(
                OzonDecision(
                    row_number=row_number,
                    entity_key_1=entity_key_1,
                    entity_key_2=entity_key_2,
                    min_price=min_price,
                    min_boost=min_boost,
                    max_boost=max_boost,
                    stock=stock,
                    final_participation=False,
                    final_promo_price=None,
                    severity=Severity.WARNING.value,
                    decision_reason="insufficient_ozon_input_data",
                    message="Недостаточно данных для принятия решения",
                )
            )
        return decisions

    def _build_summary(self, run: RunDTO, decisions: list[OzonDecision]) -> dict[str, object]:
        participating = sum(1 for decision in decisions if decision.final_participation)
        not_participating = len(decisions) - participating
        warnings = sum(1 for decision in decisions if decision.severity == Severity.WARNING.value)
        return {
            "module_code": run.module_code,
            "operation_type": run.operation_type,
            "total_rows": len(decisions),
            "processed_rows": len(decisions),
            "warnings": warnings,
            "participating": participating,
            "not_participating": not_participating,
            "no_min_price": sum(1 for decision in decisions if decision.decision_reason == "missing_min_price"),
            "no_stock": sum(1 for decision in decisions if decision.decision_reason == "no_stock"),
            "no_boosts": sum(1 for decision in decisions if decision.decision_reason == "no_boost_prices"),
            "use_max_boost_price": sum(1 for decision in decisions if decision.decision_reason == "use_max_boost_price"),
            "use_min_price": sum(1 for decision in decisions if decision.decision_reason == "use_min_price"),
            "below_min_price_threshold": sum(1 for decision in decisions if decision.decision_reason == "below_min_price_threshold"),
            "changed_rows": participating,
        }

    def _build_detail_row(self, run_id: int, decision: OzonDecision) -> RunDetailAuditDTO:
        payload = {
            "min_price": self._decimal_to_str(decision.min_price),
            "min_boost": self._decimal_to_str(decision.min_boost),
            "max_boost": self._decimal_to_str(decision.max_boost),
            "stock": self._decimal_to_str(decision.stock),
            "final_participation": decision.final_participation,
            "final_promo_price": self._decimal_to_str(decision.final_promo_price),
        }
        return RunDetailAuditDTO(
            id=0,
            run_id=run_id,
            row_number=decision.row_number,
            entity_key_1=decision.entity_key_1,
            entity_key_2=decision.entity_key_2,
            severity=decision.severity,
            decision_reason=decision.decision_reason,
            message=decision.message,
            audit_payload_json=payload,
            created_at_utc=utc_now(),
        )

    def _write_output_file(self, run: RunDTO, context: "_WorkbookContext", decisions: list[OzonDecision]) -> RunFileDTO:
        self._assert_safe_write_context(context)
        for decision in decisions:
            context.sheet.cell(row=decision.row_number, column=PARTICIPATION_COL).value = "Да" if decision.final_participation else None
            context.sheet.cell(row=decision.row_number, column=FINAL_PRICE_COL).value = (
                self._decimal_to_output(decision.final_promo_price) if decision.final_participation and decision.final_promo_price is not None else None
            )

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_path = Path(temp_file.name)
        try:
            try:
                context.workbook.save(temp_path)
            except Exception as exc:  # noqa: BLE001
                raise ValidationFailedError(
                    "Ozon workbook cannot be saved safely",
                    {"file": context.file.original_filename, "error": str(exc)},
                ) from exc
            stored = self._file_storage.copy_to_run_output(
                source_path=temp_path,
                module_code=run.module_code,
                store_id=run.store_id,
                public_run_number=run.public_run_number,
                original_filename=context.file.original_filename,
                mime_type=XLSX_MIME,
                created_at_utc=utc_now(),
            )
        finally:
            temp_path.unlink(missing_ok=True)
        expires_at = utc_now() + timedelta(days=RunFileRetentionPolicy().ttl_days)
        return RunFileDTO(
            id=0,
            run_id=run.id,
            file_role=FileRole.OZON_RESULT_OUTPUT.value,
            original_filename=context.file.original_filename,
            stored_filename=stored.stored_filename,
            storage_relative_path=stored.storage_relative_path,
            mime_type=stored.mime_type,
            file_size_bytes=stored.file_size_bytes,
            file_sha256=stored.file_sha256,
            uploaded_at_utc=stored.created_at_utc,
            expires_at_utc=expires_at,
            is_available=True,
            unavailable_reason=None,
            created_at_utc=stored.created_at_utc,
        )

    def _entity_keys(self, context: "_WorkbookContext", row_number: int) -> tuple[str | None, str | None]:
        values = [
            self._normalize_string(context.sheet.cell(row=row_number, column=1).value),
            self._normalize_string(context.sheet.cell(row=row_number, column=2).value),
            self._normalize_string(context.sheet.cell(row=row_number, column=3).value),
        ]
        primary = next((value for value in values if value), None)
        secondary = next((value for value in values if value and value != primary), None)
        return primary, secondary

    def _normalize_decimal(self, value: object) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
        if not text or text == "*":
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    def _normalize_string(self, value: object) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _decimal_to_str(self, value: Decimal | None) -> str | None:
        if value is None:
            return None
        return format(value.normalize(), "f")

    def _decimal_to_output(self, value: Decimal | None) -> int | Decimal | None:
        if value is None:
            return None
        if value == value.to_integral():
            return int(value)
        return value

    def _pick_business_result(self, operation_type: str, warnings: int) -> str:
        if operation_type == "check":
            return BusinessResult.CHECK_PASSED_WITH_WARNINGS.value if warnings else BusinessResult.CHECK_PASSED.value
        return BusinessResult.COMPLETED_WITH_WARNINGS.value if warnings else BusinessResult.COMPLETED.value

    def _validate_supported_workbook(self, file: RunFileDTO) -> None:
        suffix = Path(file.original_filename).suffix.casefold()
        if suffix != SUPPORTED_WORKBOOK_SUFFIX:
            raise ValidationFailedError(
                "Ozon workbook format is unsupported",
                {"file": file.original_filename, "expected_suffix": SUPPORTED_WORKBOOK_SUFFIX, "actual_suffix": suffix or None},
            )

    def _inspect_archive(self, file: RunFileDTO, path: Path) -> tuple[str, ...]:
        try:
            with ZipFile(path) as archive:
                return tuple(item.filename for item in archive.infolist())
        except BadZipFile as exc:
            raise ValidationFailedError("Ozon workbook cannot be opened", {"file": file.original_filename, "error": str(exc)}) from exc

    def _validate_unsafe_archive_entries(self, file: RunFileDTO, archive_entries: tuple[str, ...]) -> None:
        normalized = {entry.replace("\\", "/") for entry in archive_entries}
        if MACRO_ENTRY in normalized:
            raise ValidationFailedError("Ozon workbook macros are not supported", {"file": file.original_filename})
        if any(entry.startswith(EXTERNAL_LINKS_PREFIX) for entry in normalized):
            raise ValidationFailedError("Ozon workbook contains external links", {"file": file.original_filename})

    def _assert_safe_write_context(self, context: "_WorkbookContext") -> None:
        workbook_security = getattr(context.workbook, "security", None)
        if workbook_security is not None and any(
            bool(getattr(workbook_security, attr, False))
            for attr in ("lockStructure", "lockWindows", "lockRevision")
        ):
            raise ValidationFailedError("Ozon workbook is protected and cannot be saved safely", {"file": context.file.original_filename})
        sheet_protection = getattr(context.sheet, "protection", None)
        if sheet_protection is not None and bool(getattr(sheet_protection, "sheet", False)):
            raise ValidationFailedError("Ozon worksheet is protected and cannot be saved safely", {"file": context.file.original_filename})
        for row_number in range(4, context.sheet.max_row + 1):
            for column_number, column_name in ((PARTICIPATION_COL, "K"), (FINAL_PRICE_COL, "L")):
                cell = context.sheet.cell(row=row_number, column=column_number)
                if getattr(cell, "data_type", None) == "f":
                    raise ValidationFailedError(
                        "Ozon workbook has formulas in writable columns and cannot be saved safely",
                        {"file": context.file.original_filename, "row_number": row_number, "column_letter": column_name},
                    )


@dataclass(slots=True, frozen=True)
class _WorkbookContext:
    file: RunFileDTO
    workbook: object
    sheet: object
    headers: dict[str, int]
    archive_entries: tuple[str, ...]
