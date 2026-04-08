from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from promo.file_storage.service import FileStorageService
from promo.ozon.service import OzonExecutionStrategy
from promo.shared.contracts.runs import RunDTO, RunFileDTO
from promo.shared.contracts.stores import StoreDTO
from promo.shared.enums import MarketplaceCode, StoreStatus
from promo.shared.errors import ValidationFailedError
from promo.wb.service import WBExecutionStrategy


def _dt(value: str):
    from datetime import UTC, datetime

    return datetime.fromisoformat(value).replace(tzinfo=UTC)


def _run(module_code: str, operation_type: str = "process") -> RunDTO:
    return RunDTO(
        id=1,
        public_run_number="RUN-000001",
        store_id=1,
        initiated_by_user_id=1,
        operation_type=operation_type,
        lifecycle_status="created",
        business_result=None,
        module_code=module_code,
        input_set_signature="sig",
        started_at_utc=_dt("2026-04-07T12:00:00+00:00"),
        finished_at_utc=None,
        short_result_text=None,
        result_file_id=None,
        validation_was_auto_before_process=operation_type == "process",
        created_at_utc=_dt("2026-04-07T12:00:00+00:00"),
        updated_at_utc=_dt("2026-04-07T12:00:00+00:00"),
    )


def _wb_store() -> StoreDTO:
    return StoreDTO(
        id=1,
        name="WB",
        marketplace=MarketplaceCode.WB.value,
        status=StoreStatus.ACTIVE.value,
        wb_threshold_percent=60,
        wb_fallback_no_promo_percent=40,
        wb_fallback_over_threshold_percent=25,
        created_by_user_id=1,
        created_at_utc=_dt("2026-04-07T12:00:00+00:00"),
        updated_at_utc=_dt("2026-04-07T12:00:00+00:00"),
        archived_at_utc=None,
        archived_by_user_id=None,
    )


def _ozon_store() -> StoreDTO:
    return StoreDTO(
        id=1,
        name="Ozon",
        marketplace=MarketplaceCode.OZON.value,
        status=StoreStatus.ACTIVE.value,
        wb_threshold_percent=None,
        wb_fallback_no_promo_percent=None,
        wb_fallback_over_threshold_percent=None,
        created_by_user_id=1,
        created_at_utc=_dt("2026-04-07T12:00:00+00:00"),
        updated_at_utc=_dt("2026-04-07T12:00:00+00:00"),
        archived_at_utc=None,
        archived_by_user_id=None,
    )


def _wb_price_workbook(*, formula_in_target: bool = False, protected_sheet: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Price"
    sheet.append(["Артикул WB", "Текущая цена", "Новая скидка"])
    sheet.append(["1001", 1000, "=1+1" if formula_in_target else None])
    if protected_sheet:
        sheet.protection.sheet = True
    return _workbook_bytes(workbook)


def _wb_promo_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Promo"
    sheet.append(["Артикул WB", "Плановая цена для акции", "Загружаемая скидка для участия в акции"])
    sheet.append(["1001", 850, 20])
    return _workbook_bytes(workbook)


def _ozon_workbook(*, formula_in_k: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары и цены"
    sheet["J2"] = "Минимальная цена"
    sheet["K2"] = "Участвуем в акции"
    sheet["L2"] = "Цена для акции"
    sheet["O2"] = "Цена до скидки"
    sheet["P2"] = "Цена с max скидкой"
    sheet["R2"] = "Остаток"
    sheet["A4"] = "sku-1"
    sheet["J4"] = 700
    sheet["K4"] = "=1+1" if formula_in_k else None
    sheet["O4"] = 900
    sheet["P4"] = 750
    sheet["R4"] = 5
    return _workbook_bytes(workbook)


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _with_extra_zip_entry(source_bytes: bytes, entry_name: str, entry_body: bytes = b"unsafe") -> bytes:
    source = BytesIO(source_bytes)
    target = BytesIO()
    with ZipFile(source) as input_zip, ZipFile(target, "w", ZIP_DEFLATED) as output_zip:
        for item in input_zip.infolist():
            output_zip.writestr(item, input_zip.read(item.filename))
        output_zip.writestr(entry_name, entry_body)
    return target.getvalue()


def _store_run_file(storage: FileStorageService, original_filename: str, content: bytes, file_role: str) -> RunFileDTO:
    stored = storage.write_temp_upload(
        original_filename=original_filename,
        content=content,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        created_at_utc=_dt("2026-04-07T12:00:00+00:00"),
    )
    return RunFileDTO(
        id=1 if file_role.endswith("input") else 2,
        run_id=1,
        file_role=file_role,
        original_filename=original_filename,
        stored_filename=stored.stored_filename,
        storage_relative_path=stored.storage_relative_path,
        mime_type=stored.mime_type,
        file_size_bytes=stored.file_size_bytes,
        file_sha256=stored.file_sha256,
        uploaded_at_utc=stored.created_at_utc,
        expires_at_utc=None,
        is_available=True,
        unavailable_reason=None,
        created_at_utc=stored.created_at_utc,
    )


def test_wb_process_rejects_formula_in_writable_column(tmp_path: Path) -> None:
    storage = FileStorageService(tmp_path / "storage")
    strategy = WBExecutionStrategy(storage)
    price_file = _store_run_file(storage, "price.xlsx", _wb_price_workbook(formula_in_target=True), "wb_price_input")
    promo_file = _store_run_file(storage, "promo.xlsx", _wb_promo_workbook(), "wb_promo_input")

    with pytest.raises(ValidationFailedError, match="cannot be saved safely"):
        strategy.execute(_run("wb"), _wb_store(), (price_file, promo_file))


def test_wb_process_rejects_protected_worksheet(tmp_path: Path) -> None:
    storage = FileStorageService(tmp_path / "storage")
    strategy = WBExecutionStrategy(storage)
    price_file = _store_run_file(storage, "price.xlsx", _wb_price_workbook(protected_sheet=True), "wb_price_input")
    promo_file = _store_run_file(storage, "promo.xlsx", _wb_promo_workbook(), "wb_promo_input")

    with pytest.raises(ValidationFailedError, match="protected"):
        strategy.execute(_run("wb"), _wb_store(), (price_file, promo_file))


def test_wb_validate_rejects_external_links_archive(tmp_path: Path) -> None:
    storage = FileStorageService(tmp_path / "storage")
    strategy = WBExecutionStrategy(storage)
    price_bytes = _with_extra_zip_entry(_wb_price_workbook(), "xl/externalLinks/externalLink1.xml")
    price_file = _store_run_file(storage, "price.xlsx", price_bytes, "wb_price_input")
    promo_file = _store_run_file(storage, "promo.xlsx", _wb_promo_workbook(), "wb_promo_input")

    with pytest.raises(ValidationFailedError, match="external links"):
        strategy.validate(_run("wb", operation_type="check"), _wb_store(), (price_file, promo_file))


def test_ozon_validate_rejects_macro_workbook(tmp_path: Path) -> None:
    storage = FileStorageService(tmp_path / "storage")
    strategy = OzonExecutionStrategy(storage)
    workbook_bytes = _with_extra_zip_entry(_ozon_workbook(), "xl/vbaProject.bin")
    input_file = _store_run_file(storage, "ozon.xlsx", workbook_bytes, "ozon_input")

    with pytest.raises(ValidationFailedError, match="macros are not supported"):
        strategy.validate(_run("ozon", operation_type="check"), _ozon_store(), (input_file,))


def test_ozon_process_rejects_formula_in_writable_columns(tmp_path: Path) -> None:
    storage = FileStorageService(tmp_path / "storage")
    strategy = OzonExecutionStrategy(storage)
    input_file = _store_run_file(storage, "ozon.xlsx", _ozon_workbook(formula_in_k=True), "ozon_input")

    with pytest.raises(ValidationFailedError, match="cannot be saved safely"):
        strategy.execute(_run("ozon"), _ozon_store(), (input_file,))
