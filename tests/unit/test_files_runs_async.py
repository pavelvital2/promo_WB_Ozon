from __future__ import annotations

from io import BytesIO
from dataclasses import replace
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Callable, Generic, TypeVar

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from promo.access.contracts import AccessibleStoreDTO, SessionContextDTO
from promo.file_storage.service import FileStorageService
from promo.runs.contracts import RunServiceDependencies
from promo.runs.service import (
    CHECK_FINISHED_EVENT,
    CHECK_STARTED_EVENT,
    OLD_RESULT_REMOVED_ON_NEW_SUCCESS_EVENT,
    PROCESS_FINISHED_EVENT,
    PROCESS_STARTED_EVENT,
    InMemoryRunLockManager,
    MarketplaceRunExecutionStrategy,
    RunService,
    SkeletonRunExecutionStrategy,
)
from promo.shared.contracts.audit import RunDetailAuditDTO, RunSummaryAuditDTO
from promo.shared.contracts.files import TemporaryUploadedFileDTO
from promo.shared.contracts.logs import SystemLogDTO
from promo.shared.contracts.runs import RunDTO, RunFileDTO
from promo.shared.contracts.stores import StoreDTO
from promo.shared.contracts.users import PermissionDTO, RoleDTO, UserDTO
from promo.shared.enums import MarketplaceCode, ModuleCode, RoleCode, StoreStatus, UnavailableReason
from promo.shared.errors import ActiveRunConflictError
from promo.shared.persistence.logging import RepositoryLogger
from promo.system_maintenance.retention import (
    RUN_FILES_RETENTION_APPLIED_EVENT,
    TEMPORARY_FILES_AUTO_PURGED_EVENT,
    RunFileRetentionDependencies,
    TemporaryFileRetentionDependencies,
    expire_run_files,
    purge_temporary_files,
)
from promo.temp_files.contracts import TemporaryFileServiceDependencies, TemporaryFileUploadForm
from promo.temp_files.service import TemporaryFileService


T = TypeVar("T")


class MemoryRepository(Generic[T]):
    def __init__(self, items: list[T] | None = None) -> None:
        self._items: dict[int, T] = {}
        for item in items or []:
            self.add(item)

    def get(self, key: int) -> T | None:
        return self._items.get(key)

    def list(self) -> tuple[T, ...]:
        return tuple(self._items.values())

    def add(self, entity: T) -> T:
        self._items[getattr(entity, "id")] = entity
        return entity

    def add_many(self, entities: list[T]) -> tuple[T, ...]:
        return tuple(self.add(entity) for entity in entities)

    def update(self, entity: T) -> T:
        self._items[getattr(entity, "id")] = entity
        return entity

    def delete(self, key: int) -> None:
        self._items.pop(key, None)


def _dt(value: str) -> datetime:
    return datetime.fromisoformat(value).replace(tzinfo=UTC)


def _admin_context(store: StoreDTO) -> SessionContextDTO:
    user = UserDTO(
        id=1,
        username="admin",
        password_hash="hash",
        role_id=1,
        is_blocked=False,
        created_at_utc=_dt("2026-04-06T10:00:00+00:00"),
        updated_at_utc=_dt("2026-04-06T10:00:00+00:00"),
        last_login_at_utc=None,
    )
    role = RoleDTO(id=1, code=RoleCode.ADMIN.value, name="Администратор")
    permissions = (
        PermissionDTO(id=1, code="create_store", name="create_store"),
        PermissionDTO(id=2, code="edit_store", name="edit_store"),
    )
    accessible_store = AccessibleStoreDTO(
        id=store.id,
        name=store.name,
        marketplace=MarketplaceCode(store.marketplace),
        status=StoreStatus(store.status),
    )
    return SessionContextDTO(
        user=user,
        role=role,
        permissions=permissions,
        accessible_stores=(accessible_store,),
        is_admin=True,
        can_create_store=True,
        can_edit_store=True,
        is_blocked=False,
    )


def _workbook_bytes(builder: Callable[[Workbook], None]) -> bytes:
    workbook = Workbook()
    builder(workbook)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_wb_price_bytes() -> bytes:
    def _build(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Price"
        sheet.append(["Артикул WB", "Текущая цена", "Новая скидка"])
        sheet.append(["1001", 1000, None])
        sheet.append(["1002", 500, None])

    return _workbook_bytes(_build)


def _build_wb_promo_bytes() -> bytes:
    def _build(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Promo"
        sheet.append(["Артикул WB", "Плановая цена для акции", "Загружаемая скидка для участия в акции"])
        sheet.append(["1001", 800, 30])
        sheet.append(["1001", 700, 25])

    return _workbook_bytes(_build)


def _build_ozon_bytes() -> bytes:
    def _build(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Товары и цены"
        sheet.append(["Товары", "Действующие цены", "Участие товара в глобальной акции (МОЖНО РЕДАКТИРОВАТЬ)", "Продвижение в поиске", "Примеры цен для участия в акции", "Стоки в акции", "Статус"])
        sheet.append(["OzonID", "SKU", "Артикул", "Схема", "Название", "Категория", "Цена до скидки, RUB", "Ваша цена, RUB", "Текущая цена, RUB", "Минимальная цена, RUB", "Участие товара в акции", "Итоговая цена по акции, RUB", "Количество товаров в акции, шт", "Акционный бустинг в поиске", "Цена для минимального акционного бустинга, RUB", "Цена для максимального акционного бустинга, RUB", "Рекомендованное количество товаров на 30 дней вперёд, шт.", "Остаток на складе Ozon, шт", "Остаток на моем складе, шт"])
        sheet.append(["Только для тарифа Эконом", "Зачеркнутая цена", "Цена товара без учёта акций", "Цена с учётом акций", "Минимальная цена", "Выберите Да", "Минимальный бустинг", "Количество", "На сколько поднимем", "Чтобы получить минимальный бустинг", "", "", "", "", "", "", "", "", ""])
        sheet.append(["1", "SKU1", "article1", "", "Row4", "", "", "", "", 100, None, None, None, None, 80, 120, None, 5, None])
        sheet.append(["2", "SKU2", "article2", "", "Row5", "", "", "", "", 100, None, None, None, None, 100, 95, None, 5, None])
        sheet.append(["3", "SKU3", "article3", "", "Row6", "", "", "", "", 100, None, None, None, None, 50, 80, None, 5, None])
        sheet.append(["4", "SKU4", "article4", "", "Row7", "", "", "", "", 100, None, None, None, None, 80, 120, None, 0, None])
        sheet.append(["5", "SKU5", "article5", "", "Row8", "", "", "", "", None, None, None, None, None, 80, 120, None, 5, None])
        sheet.append(["6", "SKU6", "article6", "", "Row9", "", "", "", "", 100, None, None, None, None, None, None, None, 5, None])

    return _workbook_bytes(_build)


def _build_environment(tmp_path: Path, execution_strategy_factory: Callable[[FileStorageService], object] | None = None):
    storage = FileStorageService(tmp_path / "storage")
    stores = MemoryRepository(
        [
            StoreDTO(
                id=1,
                name="VitalEmb",
                marketplace=MarketplaceCode.WB.value,
                status=StoreStatus.ACTIVE.value,
                wb_threshold_percent=60,
                wb_fallback_no_promo_percent=40,
                wb_fallback_over_threshold_percent=25,
                created_by_user_id=1,
                created_at_utc=_dt("2026-04-06T10:00:00+00:00"),
                updated_at_utc=_dt("2026-04-06T10:00:00+00:00"),
                archived_at_utc=None,
                archived_by_user_id=None,
            ),
            StoreDTO(
                id=2,
                name="OzonShop",
                marketplace=MarketplaceCode.OZON.value,
                status=StoreStatus.ACTIVE.value,
                wb_threshold_percent=None,
                wb_fallback_no_promo_percent=None,
                wb_fallback_over_threshold_percent=None,
                created_by_user_id=1,
                created_at_utc=_dt("2026-04-06T10:00:00+00:00"),
                updated_at_utc=_dt("2026-04-06T10:00:00+00:00"),
                archived_at_utc=None,
                archived_by_user_id=None,
            )
        ]
    )
    temp_files = MemoryRepository([])
    runs = MemoryRepository([])
    run_files = MemoryRepository([])
    run_summary_audits = MemoryRepository([])
    run_detail_audits = MemoryRepository([])
    execution_strategy = execution_strategy_factory(storage) if execution_strategy_factory is not None else SkeletonRunExecutionStrategy()
    temp_service = TemporaryFileService(
        TemporaryFileServiceDependencies(temporary_files=temp_files, file_storage=storage),
        ttl_hours=24,
        clock=lambda: _dt("2026-04-06T12:00:00+00:00"),
    )
    run_service = RunService(
        RunServiceDependencies(
            runs=runs,
            run_files=run_files,
            run_summary_audits=run_summary_audits,
            run_detail_audits=run_detail_audits,
            stores=stores,
            temporary_files=temp_files,
            file_storage=storage,
        ),
        lock_manager=InMemoryRunLockManager(),
        execution_strategy=execution_strategy,
        clock=lambda: _dt("2026-04-06T12:00:00+00:00"),
    )
    return storage, stores, temp_service, run_service, temp_files, runs, run_files, run_summary_audits, run_detail_audits


def test_temp_files_active_set_and_run_execution(tmp_path: Path) -> None:
    storage, stores, temp_service, run_service, temp_files, runs, run_files, run_summary_audits, run_detail_audits = _build_environment(tmp_path)
    context = _admin_context(stores.get(1))

    uploaded = temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=b"price-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    assert uploaded.is_active_in_current_set is True
    assert storage.root_path.joinpath(uploaded.storage_relative_path).exists()
    assert temp_service.current_set_signature(context.user.id, 1, ModuleCode.WB)
    promo_uploaded = temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="promo.xlsx",
            content=b"promo-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="promo",
        ),
    )
    assert promo_uploaded.wb_file_kind == "promo"

    run = run_service.create_check_run(context, 1)
    assert run.lifecycle_status == "created"
    assert run_service.has_active_run(1, ModuleCode.WB) is True
    assert len(run_service.pending_jobs()) == 1

    polling_before = run_service.get_run_status(run.id)
    assert polling_before.lifecycle_status == "created"
    assert polling_before.is_locked is True

    processed = run_service.drain_pending_jobs()
    assert processed == 1
    polling_after = run_service.get_run_status(run.id)
    assert polling_after.lifecycle_status == "completed"
    assert polling_after.business_result == "check_passed"
    assert polling_after.is_locked is False
    assert len(run_summary_audits.list()) == 1
    assert len(run_detail_audits.list()) == 0

    page = run_service.get_run_page(run.id)
    assert page.detail_row_count == 0
    assert len(page.files) == 2
    assert page.files[0].file_role == "wb_price_input"


def test_ozon_upload_replaces_previous_active_file(tmp_path: Path) -> None:
    _, stores, temp_service, _, temp_files, _, _, _, _ = _build_environment(tmp_path)
    context = _admin_context(stores.get(1))

    first = temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.OZON,
        form=TemporaryFileUploadForm(original_filename="ozon-1.xlsx", content=b"ozon-one", mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    )
    second = temp_service.replace_file(
        first.id,
        TemporaryFileUploadForm(original_filename="ozon-2.xlsx", content=b"ozon-two", mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    )

    active = temp_service.list_active_files(context.user.id, 1, ModuleCode.OZON)
    assert active.total_items == 1
    assert active.items[0].id == second.id
    assert temp_files.get(first.id) is not None
    assert temp_files.get(first.id).is_active_in_current_set is False
    assert temp_service.current_set_signature(context.user.id, 1, ModuleCode.OZON)


def test_replace_and_delete_temp_file_clear_active_state(tmp_path: Path) -> None:
    _, stores, temp_service, _, temp_files, _, _, _, _ = _build_environment(tmp_path)
    context = _admin_context(stores.get(1))

    first = temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="wb-1.xlsx",
            content=b"wb-one",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    replaced = temp_service.replace_file(
        first.id,
        TemporaryFileUploadForm(
            original_filename="wb-2.xlsx",
            content=b"wb-two",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    assert temp_files.get(first.id) is not None
    assert temp_files.get(first.id).is_active_in_current_set is False
    assert temp_files.get(replaced.id) is not None
    assert temp_files.get(replaced.id).is_active_in_current_set is True

    temp_service.delete_file(replaced.id)
    assert temp_files.get(replaced.id) is None
    assert temp_service.list_active_files(context.user.id, 1, ModuleCode.WB).total_items == 0


def test_temp_file_actions_write_mandatory_log_events(tmp_path: Path) -> None:
    storage = FileStorageService(tmp_path / "storage")
    temp_files = MemoryRepository([])
    logs = MemoryRepository[SystemLogDTO]([])
    temp_service = TemporaryFileService(
        TemporaryFileServiceDependencies(temporary_files=temp_files, file_storage=storage),
        ttl_hours=24,
        clock=lambda: _dt("2026-04-06T12:00:00+00:00"),
        logger=RepositoryLogger(logs, clock=lambda: _dt("2026-04-06T12:00:00+00:00")),
    )

    uploaded = temp_service.upload_file(
        uploaded_by_user_id=1,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=b"price",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    replaced = temp_service.replace_file(
        uploaded.id,
        TemporaryFileUploadForm(
            original_filename="price-2.xlsx",
            content=b"price-2",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    temp_service.delete_file(replaced.id)

    event_types = [item.event_type for item in logs.list()]
    assert event_types == ["file_uploaded", "temporary_file_replaced", "temporary_file_deleted"]
    assert logs.list()[0].payload_json["file_metadata_id"] == str(uploaded.id)
    assert logs.list()[1].payload_json["replaced_file_metadata_id"] == str(uploaded.id)
    assert logs.list()[2].payload_json["file_metadata_id"] == str(replaced.id)


def test_active_run_conflict_is_blocked(tmp_path: Path) -> None:
    _, stores, temp_service, run_service, _, runs, _, _, _ = _build_environment(tmp_path)
    context = _admin_context(stores.get(1))

    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=b"price-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="promo.xlsx",
            content=b"promo-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="promo",
        ),
    )

    first_run = run_service.create_process_run(context, 1)
    assert first_run.validation_was_auto_before_process is True
    with pytest.raises(ActiveRunConflictError):
        run_service.create_check_run(context, 1)

    assert runs.get(first_run.id) is not None


def test_process_run_exposes_validating_before_completion(tmp_path: Path) -> None:
    _, stores, temp_service, run_service, _, runs, _, _, _ = _build_environment(tmp_path)
    context = _admin_context(stores.get(1))

    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=b"price-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="promo.xlsx",
            content=b"promo-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="promo",
        ),
    )
    run = run_service.create_process_run(context, 1)
    assert run.lifecycle_status == "created"

    run_service.drain_pending_jobs(limit=1)
    validating = run_service.get_run_status(run.id)
    assert validating.lifecycle_status == "validating"
    assert validating.execution_phase == "validating"
    assert validating.business_result is None
    assert len(run_service.pending_jobs()) == 1

    run_service.drain_pending_jobs(limit=1)
    completed = run_service.get_run_status(run.id)
    assert completed.lifecycle_status == "completed"
    assert completed.execution_phase == "finalized"
    assert completed.business_result == "completed"
    assert runs.get(run.id) is not None


def test_retention_hooks_mark_unavailable_and_purge_temp_files(tmp_path: Path) -> None:
    storage, stores, temp_service, run_service, temp_files, runs, run_files, _, _ = _build_environment(tmp_path)
    context = _admin_context(stores.get(1))

    temp = temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=b"price-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="promo.xlsx",
            content=b"promo-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="promo",
        ),
    )
    run = run_service.create_check_run(context, 1)
    run_service.drain_pending_jobs()

    run_file = run_files.list()[0]
    expired_run_file = replace(run_file, expires_at_utc=_dt("2026-04-01T12:00:00+00:00"))
    run_files.update(expired_run_file)
    run_retention = expire_run_files(
        RunFileRetentionDependencies(run_files=run_files, file_storage=storage),
        clock=lambda: _dt("2026-04-06T12:00:00+00:00"),
    )
    assert run_retention.affected_rows == 1
    assert RUN_FILES_RETENTION_APPLIED_EVENT in run_retention.event_types
    updated_run_file = run_files.get(run_file.id)
    assert updated_run_file is not None
    assert updated_run_file.is_available is False
    assert updated_run_file.unavailable_reason == UnavailableReason.EXPIRED.value
    assert not storage.root_path.joinpath(updated_run_file.storage_relative_path).exists()

    expired_temp = replace(temp_files.get(temp.id), expires_at_utc=_dt("2026-04-01T12:00:00+00:00"))
    temp_files.update(expired_temp)
    temp_retention = purge_temporary_files(
        TemporaryFileRetentionDependencies(temporary_files=temp_files, file_storage=storage),
        clock=lambda: _dt("2026-04-06T12:00:00+00:00"),
    )
    assert temp_retention.affected_rows == 1
    assert TEMPORARY_FILES_AUTO_PURGED_EVENT in temp_retention.event_types
    assert temp_files.get(temp.id) is None
    assert not storage.root_path.joinpath(temp.storage_relative_path).exists()


def test_run_page_reflects_superseded_result_file(tmp_path: Path) -> None:
    storage, stores, temp_service, run_service, _, runs, run_files, _, _ = _build_environment(tmp_path)
    context = _admin_context(stores.get(1))

    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=b"price-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="promo.xlsx",
            content=b"promo-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="promo",
        ),
    )
    run = run_service.create_check_run(context, 1)
    run_service.drain_pending_jobs()

    output_path = storage.root_path / "runs" / "wb" / "1" / run.public_run_number / "output"
    output_path.mkdir(parents=True, exist_ok=True)
    output_file = storage.copy_to_run_output(
        source_path=storage.root_path / run_files.list()[0].storage_relative_path,
        module_code="wb",
        store_id=1,
        public_run_number=run.public_run_number,
        original_filename="result.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    run_files.add(
        RunFileDTO(
            id=999,
            run_id=run.id,
            file_role="wb_result_output",
            original_filename="result.xlsx",
            stored_filename=output_file.stored_filename,
            storage_relative_path=output_file.storage_relative_path,
            mime_type=output_file.mime_type,
            file_size_bytes=output_file.file_size_bytes,
            file_sha256=output_file.file_sha256,
            uploaded_at_utc=_dt("2026-04-06T12:00:00+00:00"),
            expires_at_utc=None,
            is_available=True,
            unavailable_reason=None,
            created_at_utc=_dt("2026-04-06T12:00:00+00:00"),
        )
    )
    runs.update(
        replace(
            runs.get(run.id),
            result_file_id=999,
        )
    )

    superseded = run_service.supersede_run_file(999)
    assert superseded.is_available is False
    assert superseded.unavailable_reason == UnavailableReason.SUPERSEDED.value
    polling = run_service.get_run_status(run.id)
    assert polling.result_file_is_available is False
    assert polling.result_file_unavailable_reason == UnavailableReason.SUPERSEDED.value


def test_wb_check_and_process_update_only_new_discount(tmp_path: Path) -> None:
    storage, stores, temp_service, run_service, _, _, _, run_summary_audits, run_detail_audits = _build_environment(
        tmp_path,
        execution_strategy_factory=MarketplaceRunExecutionStrategy,
    )
    context = _admin_context(stores.get(1))

    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=_build_wb_price_bytes(),
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="promo.xlsx",
            content=_build_wb_promo_bytes(),
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="promo",
        ),
    )

    check_run = run_service.create_check_run(context, 1)
    run_service.drain_pending_jobs()
    check_polling = run_service.get_run_status(check_run.id)
    assert check_polling.business_result == "check_passed"
    assert check_polling.execution_phase == "finalized"
    assert len(run_summary_audits.list()) == 1
    assert len(run_detail_audits.list()) == 2

    process_run = run_service.create_process_run(context, 1)
    run_service.drain_pending_jobs(limit=1)
    validating = run_service.get_run_status(process_run.id)
    assert validating.execution_phase == "validating"
    assert validating.business_result is None
    assert len(run_service.pending_jobs()) == 1

    run_service.drain_pending_jobs(limit=1)
    completed = run_service.get_run_status(process_run.id)
    assert completed.business_result == "completed"
    assert completed.result_file_is_available is True
    assert completed.execution_phase == "finalized"

    result_file = run_service.get_run_page(process_run.id).files[-1]
    output_path = storage.root_path / result_file.storage_relative_path
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook.worksheets[0]
    assert sheet["A1"].value == "Артикул WB"
    assert sheet["B1"].value == "Текущая цена"
    assert sheet["C1"].value == "Новая скидка"
    assert sheet["A2"].value == "1001"
    assert sheet["B2"].value == 1000
    assert sheet["C2"].value == 20
    assert sheet["A3"].value == "1002"
    assert sheet["B3"].value == 500
    assert sheet["C3"].value == 40


def test_ozon_check_and_process_write_only_k_and_l(tmp_path: Path) -> None:
    storage, stores, temp_service, run_service, _, _, _, _, _ = _build_environment(
        tmp_path,
        execution_strategy_factory=MarketplaceRunExecutionStrategy,
    )
    context = _admin_context(stores.get(2))

    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=2,
        module_code=ModuleCode.OZON,
        form=TemporaryFileUploadForm(
            original_filename="ozon.xlsx",
            content=_build_ozon_bytes(),
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    )

    check_run = run_service.create_check_run(context, 2)
    run_service.drain_pending_jobs()
    check_polling = run_service.get_run_status(check_run.id)
    assert check_polling.business_result == "check_passed_with_warnings"
    assert check_polling.execution_phase == "finalized"

    process_run = run_service.create_process_run(context, 2)
    run_service.drain_pending_jobs(limit=1)
    validating = run_service.get_run_status(process_run.id)
    assert validating.execution_phase == "validating"
    assert len(run_service.pending_jobs()) == 1

    run_service.drain_pending_jobs(limit=1)
    completed = run_service.get_run_status(process_run.id)
    assert completed.business_result == "completed_with_warnings"
    assert completed.result_file_is_available is True
    assert completed.execution_phase == "finalized"

    result_file = run_service.get_run_page(process_run.id).files[-1]
    output_path = storage.root_path / result_file.storage_relative_path
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Товары и цены"]
    assert sheet.cell(row=4, column=11).value == "Да"
    assert sheet.cell(row=4, column=12).value == 120
    assert sheet.cell(row=5, column=11).value == "Да"
    assert sheet.cell(row=5, column=12).value == 100
    assert sheet.cell(row=6, column=11).value is None
    assert sheet.cell(row=6, column=12).value is None
    assert sheet.cell(row=7, column=11).value is None
    assert sheet.cell(row=7, column=12).value is None
    assert sheet.cell(row=8, column=11).value is None
    assert sheet.cell(row=8, column=12).value is None
    assert sheet.cell(row=9, column=11).value is None
    assert sheet.cell(row=9, column=12).value is None


def test_new_successful_process_supersedes_previous_result_and_writes_log(tmp_path: Path) -> None:
    storage = FileStorageService(tmp_path / "storage")
    stores = MemoryRepository(
        [
            StoreDTO(
                id=1,
                name="VitalEmb",
                marketplace=MarketplaceCode.WB.value,
                status=StoreStatus.ACTIVE.value,
                wb_threshold_percent=60,
                wb_fallback_no_promo_percent=40,
                wb_fallback_over_threshold_percent=25,
                created_by_user_id=1,
                created_at_utc=_dt("2026-04-06T10:00:00+00:00"),
                updated_at_utc=_dt("2026-04-06T10:00:00+00:00"),
                archived_at_utc=None,
                archived_by_user_id=None,
            )
        ]
    )
    temp_files = MemoryRepository([])
    runs = MemoryRepository([])
    run_files = MemoryRepository([])
    run_summary_audits = MemoryRepository([])
    run_detail_audits = MemoryRepository([])
    logs = MemoryRepository[SystemLogDTO]([])
    clock_state = {"current": _dt("2026-04-06T12:00:00+00:00")}

    def _clock() -> datetime:
        return clock_state["current"]

    temp_service = TemporaryFileService(
        TemporaryFileServiceDependencies(temporary_files=temp_files, file_storage=storage),
        ttl_hours=24,
        clock=_clock,
    )
    run_service = RunService(
        RunServiceDependencies(
            runs=runs,
            run_files=run_files,
            run_summary_audits=run_summary_audits,
            run_detail_audits=run_detail_audits,
            stores=stores,
            temporary_files=temp_files,
            file_storage=storage,
        ),
        lock_manager=InMemoryRunLockManager(),
        execution_strategy=MarketplaceRunExecutionStrategy(storage),
        clock=_clock,
        logger=RepositoryLogger(logs, clock=_clock),
    )
    context = _admin_context(stores.get(1))

    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=_build_wb_price_bytes(),
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="promo.xlsx",
            content=_build_wb_promo_bytes(),
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="promo",
        ),
    )

    first_run = run_service.create_process_run(context, 1)
    assert run_service.drain_pending_jobs(limit=1) == 1
    clock_state["current"] = _dt("2026-04-06T12:01:00+00:00")
    assert run_service.drain_pending_jobs(limit=1) == 1
    first_status = run_service.get_run_status(first_run.id)
    assert first_status.result_file_is_available is True

    clock_state["current"] = _dt("2026-04-06T12:02:00+00:00")
    second_run = run_service.create_process_run(context, 1)
    assert run_service.drain_pending_jobs(limit=1) == 1
    clock_state["current"] = _dt("2026-04-06T12:03:00+00:00")
    assert run_service.drain_pending_jobs(limit=1) == 1

    first_page = run_service.get_run_page(first_run.id)
    second_page = run_service.get_run_page(second_run.id)
    assert first_page.run.result_file_is_available is False
    assert first_page.run.result_file_unavailable_reason == UnavailableReason.SUPERSEDED.value
    assert second_page.run.result_file_is_available is True
    assert second_page.run.result_file_unavailable_reason is None
    assert len([run for run in runs.list() if run.operation_type == "process"]) == 2

    supersede_logs = [item for item in logs.list() if item.event_type == OLD_RESULT_REMOVED_ON_NEW_SUCCESS_EVENT]
    assert len(supersede_logs) == 1
    supersede_log = supersede_logs[0]
    assert supersede_log.run_id == first_run.id
    assert supersede_log.store_id == 1
    assert supersede_log.module_code == ModuleCode.WB.value
    assert supersede_log.payload_json is not None
    assert supersede_log.payload_json["superseded_run_id"] == str(first_run.id)
    assert supersede_log.payload_json["replacement_run_id"] == str(second_run.id)
    assert supersede_log.payload_json["reason"] == UnavailableReason.SUPERSEDED.value
    event_types = [item.event_type for item in logs.list()]
    assert event_types.count(PROCESS_STARTED_EVENT) == 2
    assert event_types.count(PROCESS_FINISHED_EVENT) == 2


def test_check_run_writes_started_and_finished_log_events(tmp_path: Path) -> None:
    storage, stores, temp_service, run_service, _, _, _, _, _ = _build_environment(tmp_path)
    logs = MemoryRepository[SystemLogDTO]([])
    run_service = RunService(
        RunServiceDependencies(
            runs=run_service._dependencies.runs,  # type: ignore[attr-defined]
            run_files=run_service._dependencies.run_files,  # type: ignore[attr-defined]
            run_summary_audits=run_service._dependencies.run_summary_audits,  # type: ignore[attr-defined]
            run_detail_audits=run_service._dependencies.run_detail_audits,  # type: ignore[attr-defined]
            stores=run_service._dependencies.stores,  # type: ignore[attr-defined]
            temporary_files=run_service._dependencies.temporary_files,  # type: ignore[attr-defined]
            file_storage=run_service._dependencies.file_storage,  # type: ignore[attr-defined]
        ),
        lock_manager=run_service.lock_manager(),
        execution_strategy=SkeletonRunExecutionStrategy(),
        queue=run_service._queue,  # type: ignore[attr-defined]
        clock=lambda: _dt("2026-04-06T12:00:00+00:00"),
        logger=RepositoryLogger(logs, clock=lambda: _dt("2026-04-06T12:00:00+00:00")),
    )
    context = _admin_context(stores.get(1))
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=b"price-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="promo.xlsx",
            content=b"promo-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="promo",
        ),
    )

    run = run_service.create_check_run(context, 1)
    assert run_service.drain_pending_jobs() == 1

    event_types = [item.event_type for item in logs.list()]
    assert CHECK_STARTED_EVENT in event_types
    assert CHECK_FINISHED_EVENT in event_types


def test_timeout_reconciliation_finalizes_stuck_run_and_releases_lock(tmp_path: Path) -> None:
    _, stores, temp_service, run_service, _, runs, _, run_summary_audits, _ = _build_environment(tmp_path)
    logs = MemoryRepository[SystemLogDTO]([])
    run_service = RunService(
        RunServiceDependencies(
            runs=run_service._dependencies.runs,  # type: ignore[attr-defined]
            run_files=run_service._dependencies.run_files,  # type: ignore[attr-defined]
            run_summary_audits=run_service._dependencies.run_summary_audits,  # type: ignore[attr-defined]
            run_detail_audits=run_service._dependencies.run_detail_audits,  # type: ignore[attr-defined]
            stores=run_service._dependencies.stores,  # type: ignore[attr-defined]
            temporary_files=run_service._dependencies.temporary_files,  # type: ignore[attr-defined]
            file_storage=run_service._dependencies.file_storage,  # type: ignore[attr-defined]
        ),
        lock_manager=run_service.lock_manager(),
        execution_strategy=SkeletonRunExecutionStrategy(),
        queue=run_service._queue,  # type: ignore[attr-defined]
        clock=lambda: _dt("2026-04-06T12:00:00+00:00"),
        logger=RepositoryLogger(logs, clock=lambda: _dt("2026-04-06T12:00:00+00:00")),
    )
    context = _admin_context(stores.get(1))
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="price.xlsx",
            content=b"price-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="price",
        ),
    )
    temp_service.upload_file(
        uploaded_by_user_id=context.user.id,
        store_id=1,
        module_code=ModuleCode.WB,
        form=TemporaryFileUploadForm(
            original_filename="promo.xlsx",
            content=b"promo-bytes",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            wb_file_kind="promo",
        ),
    )

    run = run_service.create_check_run(context, 1)
    timed_out_run = replace(
        runs.get(run.id),
        started_at_utc=_dt("2026-04-06T06:00:00+00:00"),
        updated_at_utc=_dt("2026-04-06T06:00:00+00:00"),
    )
    runs.update(timed_out_run)

    assert run_service.reconcile_timed_out_runs() == 1

    final_run = runs.get(run.id)
    assert final_run.lifecycle_status == "failed"
    assert final_run.business_result == "check_failed"
    assert final_run.short_result_text == "Run exceeded hard timeout"
    assert final_run.finished_at_utc == _dt("2026-04-06T12:00:00+00:00")
    assert run_service.has_active_run(1, ModuleCode.WB) is False
    assert len(run_summary_audits.list()) == 1

    timeout_logs = [item for item in logs.list() if item.event_type == "system_error"]
    assert len(timeout_logs) == 1
    assert timeout_logs[0].severity == "error"
    assert timeout_logs[0].run_id == run.id
    assert timeout_logs[0].payload_json is not None
    assert timeout_logs[0].payload_json["maintenance_task"] == "timeout_reconciliation"
    assert timeout_logs[0].payload_json["reason"] == "hard_timeout"
    assert timeout_logs[0].payload_json["timeout_seconds"] == "300"
