from __future__ import annotations

from io import BytesIO
from dataclasses import replace
from datetime import UTC, datetime

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from promo.shared.persistence.http import StoresController
from promo.shared.errors import ValidationFailedError
from promo.shared.enums import ModuleCode


def _dt(value: str) -> datetime:
    return datetime.fromisoformat(value).replace(tzinfo=UTC)


def test_startup_health_and_openapi_surface(app_harness) -> None:
    health = app_harness.client.get("/health")
    assert health.status_code == 200
    assert health.json() == {
        "status": "ok",
        "app_name": "promo-web-test",
        "environment": "test",
    }

    openapi = app_harness.client.get("/openapi.json")
    assert openapi.status_code == 200
    paths = openapi.json()["paths"]
    assert "/health" in paths
    assert "/api/auth/login" in paths
    assert "/api/auth/logout" in paths
    assert "/api/users" in paths
    assert "/api/users/{user_id}" in paths
    assert "/api/users/{user_id}/block" in paths
    assert "/api/users/{user_id}/unblock" in paths
    assert "/api/users/{user_id}/permissions/{permission_code}" in paths
    assert "/api/access/users/{user_id}/stores" in paths
    assert "/api/access/users/{user_id}/stores/{store_id}" in paths
    assert "/api/stores" in paths
    assert "/api/temp-files" in paths
    assert "/api/runs/check" in paths
    assert "/api/runs/process" in paths
    assert "/api/runs/{run_id}" in paths
    assert "/api/run-files/{run_file_id}/download" in paths
    assert "/api/history" in paths
    assert "/api/logs" in paths
    assert "/api/runs/drain" not in paths
    assert "/internal/runs/drain" not in paths
    assert "/internal/maintenance/expire-run-files" not in paths


def test_internal_ops_boundary_is_separate_from_public_surface(app_harness) -> None:
    public_runs = app_harness.client.app.state.controllers.runs
    internal_ops = app_harness.client.app.state.internal_operations

    assert not hasattr(public_runs, "drain_pending")
    assert hasattr(internal_ops.worker_runs, "drain_pending")
    assert hasattr(internal_ops.maintenance, "expire_run_files")

    missing_worker_route = app_harness.client.post("/internal/runs/drain")
    assert missing_worker_route.status_code == 404

    missing_maintenance_route = app_harness.client.post("/internal/maintenance/purge-temp-files")
    assert missing_maintenance_route.status_code == 404


def test_internal_worker_drain_contract_rejects_invalid_limit(app_harness) -> None:
    with pytest.raises(ValidationFailedError):
        app_harness.internal_drain_runs(limit=0)


def test_auth_login_logout_and_me_surface(app_harness) -> None:
    admin_token = app_harness.login("admin", "admin-pass")

    me = app_harness.client.get("/api/me", headers=app_harness.headers(admin_token))
    assert me.status_code == 200
    assert me.json()["context"]["username"] == "admin"
    assert me.json()["menu_visibility"]["show_stores"] is True

    logout = app_harness.client.post(
        "/api/auth/logout",
        headers=app_harness.headers(admin_token),
    )
    assert logout.status_code == 204

    after_logout = app_harness.client.get("/api/me", headers=app_harness.headers(admin_token))
    assert after_logout.status_code == 403


def test_ui_login_and_dashboard_routes(app_harness) -> None:
    login_page = app_harness.client.get("/login")
    assert login_page.status_code == 200
    assert "Вход" in login_page.text
    assert "username" in login_page.text.lower()
    assert "password" in login_page.text.lower()

    admin_token = app_harness.login("admin", "admin-pass")
    app_harness.use_ui_session(admin_token)

    dashboard = app_harness.client.get("/dashboard")
    assert dashboard.status_code == 200
    assert "Dashboard" in dashboard.text
    assert "/users" in dashboard.text
    assert "/stores" in dashboard.text
    assert "/processing/wb" in dashboard.text
    assert "/processing/ozon" in dashboard.text
    assert "/runs" in dashboard.text
    assert "/logs" in dashboard.text

    users_page = app_harness.client.get("/users")
    assert users_page.status_code == 200
    assert "Пользователи" in users_page.text
    assert "Created" in users_page.text
    assert "Last login" in users_page.text
    assert "Заблокировать" in users_page.text

    create_user_page = app_harness.client.get("/users/create")
    assert create_user_page.status_code == 200
    assert "Создать пользователя" in create_user_page.text

    profile_password = app_harness.client.get("/profile/password")
    assert profile_password.status_code == 200
    assert "Смена пароля" in profile_password.text
    assert "confirm_new_password" in profile_password.text


def test_ui_hidden_sections_and_no_store_dashboard(app_harness) -> None:
    admin_token = app_harness.login("admin", "admin-pass")

    create_user = app_harness.client.post(
        "/api/users",
        headers=app_harness.headers(admin_token),
        json={
            "username": "lead-ui",
            "password": "lead-ui-pass",
            "role_code": "manager_lead",
            "permission_codes": ["create_store"],
        },
    )
    assert create_user.status_code == 200

    manager_token = app_harness.login("manager", "manager-pass")
    app_harness.use_ui_session(manager_token)
    manager_dashboard = app_harness.client.get("/dashboard")
    assert manager_dashboard.status_code == 200
    assert "Нет доступных магазинов" in manager_dashboard.text
    assert 'href="/users"' not in manager_dashboard.text
    assert 'href="/logs"' not in manager_dashboard.text
    assert 'href="/stores"' not in manager_dashboard.text

    denied_history = app_harness.client.get("/runs")
    assert denied_history.status_code == 403

    lead_token = app_harness.login("lead-ui", "lead-ui-pass")
    app_harness.use_ui_session(lead_token)
    lead_dashboard = app_harness.client.get("/dashboard")
    assert lead_dashboard.status_code == 200
    assert "Нет доступных магазинов" in lead_dashboard.text
    assert 'href="/stores"' in lead_dashboard.text
    assert 'href="/stores/create"' in lead_dashboard.text
    assert 'href="/users"' not in lead_dashboard.text
    assert 'href="/logs"' not in lead_dashboard.text

    stores_create = app_harness.client.get("/stores/create")
    assert stores_create.status_code == 200
    assert "Создать магазин" in stores_create.text


def test_ui_pages_cover_history_run_page_processing_and_logs(app_harness) -> None:
    admin_token = app_harness.login("admin", "admin-pass")
    manager_token = app_harness.login("manager", "manager-pass")
    managed_user = app_harness.client.post(
        "/api/users",
        headers=app_harness.headers(admin_token),
        json={
            "username": "store-ui-user",
            "password": "store-ui-pass",
            "role_code": "manager",
            "permission_codes": [],
        },
    )
    assert managed_user.status_code == 200

    store = app_harness.create_store(
        admin_token,
        name="UiOzonStore",
        marketplace="ozon",
    )
    app_harness.grant_store_access(admin_token, user_id=2, store_id=store["id"])
    uploaded = app_harness.upload_temp_file(
        manager_token,
        store_id=store["id"],
        module_code=ModuleCode.OZON,
        original_filename="ozon.xlsx",
        content=b"xlsx-ozon",
    )
    assert uploaded.status_code == 200

    run = app_harness.client.post(
        "/api/runs/check",
        headers=app_harness.headers(manager_token),
        json={"store_id": store["id"]},
    )
    assert run.status_code == 200
    app_harness.internal_drain_runs()

    app_harness.use_ui_session(manager_token)
    processing = app_harness.client.get(f"/processing/ozon?store_id={store['id']}&run_public={run.json()['public_run_number']}")
    assert processing.status_code == 200
    assert "Ozon Processing" in processing.text
    assert "Активный временный набор" in processing.text
    assert "Последний запуск" in processing.text
    assert "Заменить" in processing.text

    wb_store = app_harness.create_store(
        admin_token,
        name="UiWbStorePage",
        marketplace="wb",
        wb_threshold_percent=60,
        wb_fallback_no_promo_percent=40,
        wb_fallback_over_threshold_percent=25,
    )
    app_harness.grant_store_access(admin_token, user_id=2, store_id=wb_store["id"])
    processing_wb = app_harness.client.get(f"/processing/wb?store_id={wb_store['id']}")
    assert processing_wb.status_code == 200
    assert "Wildberries Processing" in processing_wb.text
    assert "WB file kind" in processing_wb.text

    history = app_harness.client.get("/runs?store_id={}&sort_field=public_run_number&descending=false".format(store["id"]))
    assert history.status_code == 200
    assert run.json()["public_run_number"] in history.text
    assert "name=\"store_id\"" in history.text
    assert "name=\"sort_field\"" in history.text
    assert "name=\"operation_type\"" in history.text
    assert "Предыдущая" in history.text
    assert "Следующая" in history.text

    run_page = app_harness.client.get(f"/runs/{run.json()['public_run_number']}?severity=info&sort_field=row_number&page_size=25")
    assert run_page.status_code == 200
    assert "Summary audit" in run_page.text
    assert "Detail audit" in run_page.text
    assert store["name"] in run_page.text
    assert "name=\"severity\"" in run_page.text
    assert "name=\"decision_reason\"" in run_page.text
    assert "name=\"row_number_from\"" in run_page.text
    assert "Предыдущая" in run_page.text
    assert "Следующая" in run_page.text

    app_harness.use_ui_session(admin_token)
    store_edit = app_harness.client.get(f"/stores/{store['id']}/edit")
    assert store_edit.status_code == 200
    assert "Назначенные пользователи" in store_edit.text
    assert "Добавить пользователя" in store_edit.text
    assert f"/api/access/users/{managed_user.json()['id']}/stores/{store['id']}" in store_edit.text

    user_edit = app_harness.client.get(f"/users/{managed_user.json()['id']}/edit")
    assert user_edit.status_code == 200
    assert "Permissions" in user_edit.text
    assert "Назначить permission" in user_edit.text

    admin_logs = app_harness.login("admin", "admin-pass")
    app_harness.use_ui_session(admin_logs)
    logs = app_harness.client.get("/logs?event_type=successful_login&sort_field=event_type&descending=false")
    assert logs.status_code == 200
    assert "Logs" in logs.text
    assert "successful_login" in logs.text
    assert "name=\"event_type\"" in logs.text
    assert "name=\"severity\"" in logs.text
    assert "name=\"public_run_number\"" in logs.text
    assert "Предыдущая" in logs.text
    assert "Следующая" in logs.text

    app_harness.use_ui_session(manager_token)
    forbidden_logs = app_harness.client.get("/logs")
    assert forbidden_logs.status_code == 403


def test_autonomous_runtime_completes_run_via_public_surface_without_manual_drain(autonomous_app_harness) -> None:
    admin_token = autonomous_app_harness.login("admin", "admin-pass")
    manager_token = autonomous_app_harness.login("manager", "manager-pass")

    store = autonomous_app_harness.create_store(
        admin_token,
        name="AutonomousSurfaceOzon",
        marketplace="ozon",
    )
    autonomous_app_harness.grant_store_access(admin_token, user_id=2, store_id=store["id"])

    uploaded = autonomous_app_harness.upload_temp_file(
        manager_token,
        store_id=store["id"],
        module_code=ModuleCode.OZON,
        original_filename="ozon.xlsx",
        content=b"ozon-bytes",
    )
    assert uploaded.status_code == 200

    run = autonomous_app_harness.client.post(
        "/api/runs/check",
        headers=autonomous_app_harness.headers(manager_token),
        json={"store_id": store["id"]},
    )
    assert run.status_code == 200

    status_payload = autonomous_app_harness.wait_for_run_completion(manager_token, run.json()["id"])
    assert status_payload["lifecycle_status"] == "completed"
    assert status_payload["business_result"] == "check_passed"
    assert autonomous_app_harness.client.app.state.run_worker_runtime.is_running() is True


def test_autonomous_maintenance_runtime_reconciles_timeout_through_public_surface(autonomous_maintenance_app_harness) -> None:
    admin_token = autonomous_maintenance_app_harness.login("admin", "admin-pass")
    manager_token = autonomous_maintenance_app_harness.login("manager", "manager-pass")

    store = autonomous_maintenance_app_harness.create_store(
        admin_token,
        name="AutonomousMaintenanceOzon",
        marketplace="ozon",
    )
    autonomous_maintenance_app_harness.grant_store_access(admin_token, user_id=2, store_id=store["id"])

    uploaded = autonomous_maintenance_app_harness.upload_temp_file(
        manager_token,
        store_id=store["id"],
        module_code=ModuleCode.OZON,
        original_filename="ozon.xlsx",
        content=b"ozon-bytes",
    )
    assert uploaded.status_code == 200

    run = autonomous_maintenance_app_harness.client.post(
        "/api/runs/check",
        headers=autonomous_maintenance_app_harness.headers(manager_token),
        json={"store_id": store["id"]},
    )
    assert run.status_code == 200
    run_id = run.json()["id"]

    with autonomous_maintenance_app_harness.app_context.request_scope(commit=True) as bundle:
        run_record = bundle.uow.repositories.runs.get(run_id)
        assert run_record is not None
        bundle.uow.repositories.runs.update(
            replace(
                run_record,
                started_at_utc=_dt("2026-04-07T06:00:00+00:00"),
                updated_at_utc=_dt("2026-04-07T06:00:00+00:00"),
            )
        )

    status_payload = autonomous_maintenance_app_harness.wait_for_run_completion(manager_token, run_id)
    assert status_payload["lifecycle_status"] == "failed"
    assert status_payload["business_result"] == "check_failed"
    assert status_payload["is_locked"] is False

    logs = autonomous_maintenance_app_harness.client.get(
        "/api/logs?page=1&page_size=25&event_type=system_error&search=%s" % run.json()["public_run_number"],
        headers=autonomous_maintenance_app_harness.headers(admin_token),
    )
    assert logs.status_code == 200
    assert any(item["payload_json"]["maintenance_task"] == "timeout_reconciliation" for item in logs.json()["items"])
    assert autonomous_maintenance_app_harness.client.app.state.maintenance_runtime.is_running() is True


def test_structured_errors_are_normalized_on_public_boundaries(app_harness) -> None:
    missing_session = app_harness.client.get("/api/me")
    assert missing_session.status_code == 401
    assert missing_session.json()["error_code"] == "access_denied"
    assert missing_session.json()["details"]["auth_error"] == "missing_session_token"

    admin_token = app_harness.login("admin", "admin-pass")
    invalid_base64 = app_harness.client.post(
        "/api/temp-files?store_id=1&module_code=ozon",
        headers=app_harness.headers(admin_token),
        json={
            "original_filename": "broken.xlsx",
            "content_base64": "!!!",
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )
    assert invalid_base64.status_code == 400
    assert invalid_base64.json()["error_code"] == "validation_failed"
    assert invalid_base64.json()["details"]["validation_error"] == "invalid_base64"

    validation_422 = app_harness.client.get(
        "/api/history?page=bad&page_size=25",
        headers=app_harness.headers(admin_token),
    )
    assert validation_422.status_code == 422
    assert validation_422.json()["error_code"] == "validation_failed"
    assert validation_422.json()["details"]["validation_error"] == "request_validation_failed"
    assert validation_422.json()["details"]["validation_errors"]


def test_unexpected_exception_is_normalized_and_persisted(app_harness, monkeypatch) -> None:
    admin_token = app_harness.login("admin", "admin-pass")

    def _boom(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(StoresController, "list_stores", _boom)
    failing_client = TestClient(app_harness.client.app, raise_server_exceptions=False)
    response = failing_client.get("/api/stores", headers=app_harness.headers(admin_token))

    assert response.status_code == 500
    payload = response.json()
    assert payload["error_code"] == "system_error"
    assert payload["error_message"] == "Unexpected system error"
    assert payload["details"]["exception_type"] == "RuntimeError"
    assert payload["details"]["request_path"] == "/api/stores"

    with app_harness.app_context.request_scope() as bundle:
        system_errors = [item for item in bundle.uow.repositories.logs.list() if item.event_type == "system_error"]
    assert len(system_errors) == 1
    assert system_errors[0].severity == "error"
    assert system_errors[0].payload_json is not None
    assert system_errors[0].payload_json["exception_type"] == "RuntimeError"
    assert system_errors[0].payload_json["request_path"] == "/api/stores"


def test_blocked_user_denial_and_no_store_behavior(app_harness) -> None:
    blocked_login = app_harness.login_response("blocked-manager", "blocked-pass")
    assert blocked_login.status_code == 403

    manager_token = app_harness.login("manager", "manager-pass")
    me = app_harness.client.get("/api/me", headers=app_harness.headers(manager_token))
    assert me.status_code == 200
    assert me.json()["context"]["username"] == "manager"
    assert me.json()["context"]["accessible_store_count"] == 0
    assert me.json()["menu_visibility"]["show_no_store_state"] is True
    assert me.json()["no_store_state"]["message"] == "Нет доступных магазинов"
    assert me.json()["no_store_state"]["can_create_store"] is False

    history = app_harness.client.get("/api/history", headers=app_harness.headers(manager_token))
    assert history.status_code == 403


def test_change_password_smoke(app_harness) -> None:
    admin_token = app_harness.login("admin", "admin-pass")

    change = app_harness.client.post(
        "/api/auth/change-password",
        headers=app_harness.headers(admin_token),
        json={
            "current_password": "admin-pass",
            "new_password": "admin-pass-2",
        },
    )
    assert change.status_code == 200
    assert change.json()["user_id"] == 1

    old_login = app_harness.login_response("admin", "admin-pass")
    assert old_login.status_code == 403

    new_token = app_harness.login("admin", "admin-pass-2")
    me = app_harness.client.get("/api/me", headers=app_harness.headers(new_token))
    assert me.status_code == 200
    assert me.json()["context"]["username"] == "admin"


def test_end_to_end_happy_path_acceptance_via_http_access_surface(app_harness) -> None:
    admin_token = app_harness.login("admin", "admin-pass")
    manager_token = app_harness.login("manager", "manager-pass")

    store = app_harness.create_store(
        admin_token,
        name="AcceptanceOzon",
        marketplace="ozon",
    )
    store_id = store["id"]
    app_harness.grant_store_access(admin_token, user_id=2, store_id=store_id)

    access_list = app_harness.client.get(
        "/api/access/users/2/stores",
        headers=app_harness.headers(admin_token),
    )
    assert access_list.status_code == 200
    assert any(item["store_id"] == store_id for item in access_list.json())

    forbidden_list = app_harness.client.get(
        "/api/access/users/2/stores",
        headers=app_harness.headers(manager_token),
    )
    assert forbidden_list.status_code == 403

    uploaded = app_harness.upload_temp_file(
        manager_token,
        store_id=store_id,
        module_code=ModuleCode.OZON,
        original_filename="ozon.xlsx",
        content=b"ozon-bytes",
    )
    assert uploaded.status_code == 200
    uploaded_json = uploaded.json()
    assert uploaded_json["original_filename"] == "ozon.xlsx"

    active_files = app_harness.client.get(
        f"/api/temp-files?store_id={store_id}&module_code=ozon",
        headers=app_harness.headers(manager_token),
    )
    assert active_files.status_code == 200
    assert active_files.json()["total_items"] == 1

    run = app_harness.client.post(
        "/api/runs/check",
        headers=app_harness.headers(manager_token),
        json={"store_id": store_id},
    )
    assert run.status_code == 200
    run_id = run.json()["id"]

    assert app_harness.internal_drain_runs() == 1

    status_response = app_harness.client.get(
        f"/api/runs/{run_id}/status",
        headers=app_harness.headers(manager_token),
    )
    assert status_response.status_code == 200
    assert status_response.json()["lifecycle_status"] == "completed"
    assert status_response.json()["business_result"] == "check_passed"

    run_page = app_harness.client.get(
        f"/api/runs/{run_id}",
        headers=app_harness.headers(manager_token),
    )
    assert run_page.status_code == 200
    run_page_json = run_page.json()
    assert run_page_json["run"]["run_id"] == run_id
    assert run_page_json["run"]["store_name"] == "AcceptanceOzon"
    assert run_page_json["polling"]["id"] == run_id
    assert run_page_json["summary_audit_json"] is not None

    detail = app_harness.client.get(
        f"/api/audit/runs/{run_id}/detail?page=1&page_size=25",
        headers=app_harness.headers(manager_token),
    )
    assert detail.status_code == 200
    assert detail.json()["page"] == 1

    history = app_harness.client.get(
        "/api/history",
        headers=app_harness.headers(manager_token),
    )
    assert history.status_code == 200
    assert history.json()["total_items"] == 1
    assert history.json()["items"][0]["public_run_number"] == run.json()["public_run_number"]

    logs = app_harness.client.get(
        "/api/logs",
        headers=app_harness.headers(admin_token),
    )
    assert logs.status_code == 200
    assert logs.json()["total_items"] >= 1
    event_types = {item["event_type"] for item in logs.json()["items"]}
    assert "successful_login" in event_types


def test_admin_user_management_backend_flow_via_http_surface(app_harness) -> None:
    admin_token = app_harness.login("admin", "admin-pass")
    manager_token = app_harness.login("manager", "manager-pass")

    create_user = app_harness.client.post(
        "/api/users",
        headers=app_harness.headers(admin_token),
        json={
            "username": "managed-user",
            "password": "managed-pass",
            "role_code": "manager",
            "permission_codes": ["create_store"],
        },
    )
    assert create_user.status_code == 200
    created_user = create_user.json()
    user_id = created_user["id"]
    assert created_user["role_code"] == "manager"
    assert created_user["permission_codes"] == ["create_store"]

    edited = app_harness.client.patch(
        f"/api/users/{user_id}",
        headers=app_harness.headers(admin_token),
        json={"username": "managed-user-2", "role_code": "manager_lead"},
    )
    assert edited.status_code == 200
    assert edited.json()["role_code"] == "manager_lead"

    permission_grant = app_harness.client.post(
        f"/api/users/{user_id}/permissions/edit_store",
        headers=app_harness.headers(admin_token),
    )
    assert permission_grant.status_code == 200
    assert set(permission_grant.json()["permission_codes"]) == {"create_store", "edit_store"}

    permission_revoke = app_harness.client.delete(
        f"/api/users/{user_id}/permissions/create_store",
        headers=app_harness.headers(admin_token),
    )
    assert permission_revoke.status_code == 200
    assert permission_revoke.json()["permission_codes"] == ["edit_store"]

    blocked = app_harness.client.post(
        f"/api/users/{user_id}/block",
        headers=app_harness.headers(admin_token),
    )
    assert blocked.status_code == 200
    assert blocked.json()["is_blocked"] is True

    unblocked = app_harness.client.post(
        f"/api/users/{user_id}/unblock",
        headers=app_harness.headers(admin_token),
    )
    assert unblocked.status_code == 200
    assert unblocked.json()["is_blocked"] is False

    store = app_harness.create_store(
        admin_token,
        name="UserMgmtStore",
        marketplace="wb",
        wb_threshold_percent=60,
        wb_fallback_no_promo_percent=40,
        wb_fallback_over_threshold_percent=25,
    )
    access_grant = app_harness.client.post(
        f"/api/access/users/{user_id}/stores/{store['id']}",
        headers=app_harness.headers(admin_token),
    )
    assert access_grant.status_code == 200

    user_detail = app_harness.client.get(
        f"/api/users/{user_id}",
        headers=app_harness.headers(admin_token),
    )
    assert user_detail.status_code == 200
    assert user_detail.json()["accessible_store_count"] == 1
    assert user_detail.json()["store_access"][0]["store_id"] == store["id"]

    access_revoke = app_harness.client.delete(
        f"/api/access/users/{user_id}/stores/{store['id']}",
        headers=app_harness.headers(admin_token),
    )
    assert access_revoke.status_code == 204

    users_list = app_harness.client.get(
        "/api/users",
        headers=app_harness.headers(admin_token),
    )
    assert users_list.status_code == 200
    assert any(item["id"] == user_id for item in users_list.json()["items"])

    assert app_harness.client.get("/api/users", headers=app_harness.headers(manager_token)).status_code == 403
    assert app_harness.client.post(
        "/api/users",
        headers=app_harness.headers(manager_token),
        json={"username": "denied", "password": "denied-pass", "role_code": "manager", "permission_codes": []},
    ).status_code == 403
    assert app_harness.client.patch(
        f"/api/users/{user_id}",
        headers=app_harness.headers(manager_token),
        json={"username": "denied-edit"},
    ).status_code == 403
    assert app_harness.client.post(
        f"/api/users/{user_id}/block",
        headers=app_harness.headers(manager_token),
    ).status_code == 403
    assert app_harness.client.post(
        f"/api/users/{user_id}/permissions/create_store",
        headers=app_harness.headers(manager_token),
    ).status_code == 403
    assert app_harness.client.post(
        f"/api/access/users/{user_id}/stores/{store['id']}",
        headers=app_harness.headers(manager_token),
    ).status_code == 403

    logs = app_harness.client.get(
        "/api/logs?page=1&page_size=100",
        headers=app_harness.headers(admin_token),
    )
    assert logs.status_code == 200
    event_types = {item["event_type"] for item in logs.json()["items"]}
    assert "user_created" in event_types
    assert "user_updated" in event_types
    assert "user_blocked" in event_types
    assert "user_unblocked" in event_types


def test_negative_access_and_conflict_scenarios(app_harness) -> None:
    admin_token = app_harness.login("admin", "admin-pass")
    manager_token = app_harness.login("manager", "manager-pass")

    unauthorized_me = app_harness.client.get("/api/me")
    assert unauthorized_me.status_code == 401

    foreign_store = app_harness.create_store(
        admin_token,
        name="ForeignWB",
        marketplace="wb",
        wb_threshold_percent=60,
        wb_fallback_no_promo_percent=40,
        wb_fallback_over_threshold_percent=25,
    )
    foreign_store_id = foreign_store["id"]

    blocked_upload = app_harness.upload_temp_file(
        manager_token,
        store_id=foreign_store_id,
        module_code=ModuleCode.WB,
        original_filename="blocked.xlsx",
        content=b"blocked",
        wb_file_kind="price",
    )
    assert blocked_upload.status_code == 403

    app_harness.grant_store_access(admin_token, user_id=2, store_id=foreign_store_id)

    manager_uploaded = app_harness.upload_temp_file(
        manager_token,
        store_id=foreign_store_id,
        module_code=ModuleCode.WB,
        original_filename="allowed.xlsx",
        content=b"allowed",
        wb_file_kind="price",
    )
    assert manager_uploaded.status_code == 200
    manager_promo_uploaded = app_harness.upload_temp_file(
        manager_token,
        store_id=foreign_store_id,
        module_code=ModuleCode.WB,
        original_filename="promo.xlsx",
        content=b"promo",
        wb_file_kind="promo",
    )
    assert manager_promo_uploaded.status_code == 200

    admin_uploaded = app_harness.upload_temp_file(
        admin_token,
        store_id=foreign_store_id,
        module_code=ModuleCode.WB,
        original_filename="admin.xlsx",
        content=b"admin",
        wb_file_kind="price",
    )
    assert admin_uploaded.status_code == 200

    blocked_delete = app_harness.client.delete(
        f"/api/temp-files/{admin_uploaded.json()['id']}",
        headers=app_harness.headers(manager_token),
    )
    assert blocked_delete.status_code == 403

    logs_for_manager = app_harness.client.get(
        "/api/logs",
        headers=app_harness.headers(manager_token),
    )
    assert logs_for_manager.status_code == 403

    first_run = app_harness.client.post(
        "/api/runs/check",
        headers=app_harness.headers(manager_token),
        json={"store_id": foreign_store_id},
    )
    assert first_run.status_code == 200

    second_run = app_harness.client.post(
        "/api/runs/check",
        headers=app_harness.headers(manager_token),
        json={"store_id": foreign_store_id},
    )
    assert second_run.status_code == 409

    admin_only_store = app_harness.create_store(
        admin_token,
        name="AdminOnlyOzon",
        marketplace="ozon",
    )
    admin_only_store_id = admin_only_store["id"]

    admin_only_upload = app_harness.upload_temp_file(
        admin_token,
        store_id=admin_only_store_id,
        module_code=ModuleCode.OZON,
        original_filename="admin-only.xlsx",
        content=b"admin-only",
    )
    assert admin_only_upload.status_code == 200

    foreign_run = app_harness.client.post(
        "/api/runs/check",
        headers=app_harness.headers(admin_token),
        json={"store_id": admin_only_store_id},
    )
    assert foreign_run.status_code == 200
    foreign_run_id = foreign_run.json()["id"]

    assert app_harness.internal_drain_runs() == 2

    foreign_status = app_harness.client.get(
        f"/api/runs/{foreign_run_id}/status",
        headers=app_harness.headers(manager_token),
    )
    assert foreign_status.status_code == 403

    scoped_history = app_harness.client.get(
        "/api/history",
        headers=app_harness.headers(manager_token),
    )
    assert scoped_history.status_code == 200
    assert scoped_history.json()["total_items"] == 1
    assert scoped_history.json()["items"][0]["store_id"] == foreign_store_id


def test_archived_store_behavior_for_new_run_surface(app_harness) -> None:
    admin_token = app_harness.login("admin", "admin-pass")
    manager_token = app_harness.login("manager", "manager-pass")

    store = app_harness.create_store(
        admin_token,
        name="ArchivedWB",
        marketplace="wb",
        wb_threshold_percent=60,
        wb_fallback_no_promo_percent=40,
        wb_fallback_over_threshold_percent=25,
    )
    store_id = store["id"]
    app_harness.grant_store_access(admin_token, user_id=2, store_id=store_id)

    archive = app_harness.client.post(
        f"/api/stores/{store_id}/archive",
        headers=app_harness.headers(admin_token),
    )
    assert archive.status_code == 200
    assert archive.json()["status"] == "archived"

    stores = app_harness.client.get(
        "/api/stores",
        headers=app_harness.headers(manager_token),
    )
    assert stores.status_code == 200
    archived_item = next(item for item in stores.json()["items"] if item["id"] == store_id)
    assert archived_item["status"] == "archived"

    run = app_harness.client.post(
        "/api/runs/check",
        headers=app_harness.headers(manager_token),
        json={"store_id": store_id},
    )
    assert run.status_code == 400


def test_temp_file_intake_limits_and_composition_are_enforced_via_http_surface(app_harness, monkeypatch) -> None:
    from promo.temp_files import service as temp_file_service_module

    monkeypatch.setattr(temp_file_service_module, "MAX_FILE_SIZE_BYTES", 8)
    monkeypatch.setattr(temp_file_service_module, "MAX_WB_TOTAL_SIZE_BYTES", 12)
    monkeypatch.setattr(temp_file_service_module, "MAX_WB_PROMO_FILES", 2)

    admin_token = app_harness.login("admin", "admin-pass")
    manager_token = app_harness.login("manager", "manager-pass")

    wb_store = app_harness.create_store(
        admin_token,
        name="LimitedWB",
        marketplace="wb",
        wb_threshold_percent=60,
        wb_fallback_no_promo_percent=40,
        wb_fallback_over_threshold_percent=25,
    )
    app_harness.grant_store_access(admin_token, user_id=2, store_id=wb_store["id"])

    missing_kind = app_harness.upload_temp_file(
        manager_token,
        store_id=wb_store["id"],
        module_code=ModuleCode.WB,
        original_filename="wb.xlsx",
        content=b"1234",
    )
    assert missing_kind.status_code == 400
    assert missing_kind.json()["error_code"] == "validation_failed"

    oversized_file = app_harness.upload_temp_file(
        manager_token,
        store_id=wb_store["id"],
        module_code=ModuleCode.WB,
        original_filename="price.xlsx",
        content=b"123456789",
        wb_file_kind="price",
    )
    assert oversized_file.status_code == 400
    assert oversized_file.json()["error_code"] == "file_limit_exceeded"

    assert app_harness.upload_temp_file(
        manager_token,
        store_id=wb_store["id"],
        module_code=ModuleCode.WB,
        original_filename="price.xlsx",
        content=b"1111",
        wb_file_kind="price",
    ).status_code == 200
    assert app_harness.upload_temp_file(
        manager_token,
        store_id=wb_store["id"],
        module_code=ModuleCode.WB,
        original_filename="promo-1.xlsx",
        content=b"1111",
        wb_file_kind="promo",
    ).status_code == 200
    assert app_harness.upload_temp_file(
        manager_token,
        store_id=wb_store["id"],
        module_code=ModuleCode.WB,
        original_filename="promo-2.xlsx",
        content=b"1111",
        wb_file_kind="promo",
    ).status_code == 200

    wb_total_limit = app_harness.upload_temp_file(
        manager_token,
        store_id=wb_store["id"],
        module_code=ModuleCode.WB,
        original_filename="promo-3.xlsx",
        content=b"1",
        wb_file_kind="promo",
    )
    assert wb_total_limit.status_code == 400
    assert wb_total_limit.json()["error_code"] == "file_limit_exceeded"
    assert wb_total_limit.json()["details"]["limit_type"] == "wb_total_size"

    monkeypatch.setattr(temp_file_service_module, "MAX_WB_TOTAL_SIZE_BYTES", 20)
    wb_promo_count_limit = app_harness.upload_temp_file(
        manager_token,
        store_id=wb_store["id"],
        module_code=ModuleCode.WB,
        original_filename="promo-3.xlsx",
        content=b"11",
        wb_file_kind="promo",
    )
    assert wb_promo_count_limit.status_code == 400
    assert wb_promo_count_limit.json()["error_code"] == "file_limit_exceeded"
    assert wb_promo_count_limit.json()["details"]["limit_type"] == "wb_promo_count"

    ozon_store = app_harness.create_store(
        admin_token,
        name="EmptyOzon",
        marketplace="ozon",
    )
    app_harness.grant_store_access(admin_token, user_id=2, store_id=ozon_store["id"])

    empty_ozon_run = app_harness.client.post(
        "/api/runs/check",
        headers=app_harness.headers(manager_token),
        json={"store_id": ozon_store["id"]},
    )
    assert empty_ozon_run.status_code == 400
    assert empty_ozon_run.json()["error_code"] == "file_limit_exceeded"
    assert empty_ozon_run.json()["details"]["required_file_count"] == 1


def test_wb_check_and_process_with_download_and_unavailable_behavior(
    marketplace_app_harness,
    wb_price_bytes: bytes,
    wb_promo_bytes: bytes,
) -> None:
    admin_token = marketplace_app_harness.login("admin", "admin-pass")
    manager_token = marketplace_app_harness.login("manager", "manager-pass")

    store = marketplace_app_harness.create_store(
        admin_token,
        name="AcceptanceWB",
        marketplace="wb",
        wb_threshold_percent=60,
        wb_fallback_no_promo_percent=40,
        wb_fallback_over_threshold_percent=25,
    )
    store_id = store["id"]
    marketplace_app_harness.grant_store_access(admin_token, user_id=2, store_id=store_id)

    upload_price = marketplace_app_harness.upload_temp_file(
        manager_token,
        store_id=store_id,
        module_code=ModuleCode.WB,
        original_filename="price.xlsx",
        content=wb_price_bytes,
        wb_file_kind="price",
    )
    upload_promo = marketplace_app_harness.upload_temp_file(
        manager_token,
        store_id=store_id,
        module_code=ModuleCode.WB,
        original_filename="promo.xlsx",
        content=wb_promo_bytes,
        wb_file_kind="promo",
    )
    assert upload_price.status_code == 200
    assert upload_promo.status_code == 200

    check_run = marketplace_app_harness.client.post(
        "/api/runs/check",
        headers=marketplace_app_harness.headers(manager_token),
        json={"store_id": store_id},
    )
    assert check_run.status_code == 200
    assert marketplace_app_harness.internal_drain_runs() == 1

    check_status = marketplace_app_harness.client.get(
        f"/api/runs/{check_run.json()['id']}/status",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert check_status.status_code == 200
    assert check_status.json()["business_result"] == "check_passed"

    process_run = marketplace_app_harness.client.post(
        "/api/runs/process",
        headers=marketplace_app_harness.headers(manager_token),
        json={"store_id": store_id},
    )
    assert process_run.status_code == 200
    assert marketplace_app_harness.internal_drain_runs() == 2

    process_run_id = process_run.json()["id"]
    run_page = marketplace_app_harness.client.get(
        f"/api/runs/{process_run_id}",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert run_page.status_code == 200
    run_page_json = run_page.json()
    assert run_page_json["run"]["business_result"] == "completed"
    assert run_page_json["run"]["result_file_is_available"] is True
    assert len(run_page_json["files"]) == 3

    source_file = next(item for item in run_page_json["files"] if item["file_role"] == "wb_price_input")
    result_file = next(item for item in run_page_json["files"] if item["file_role"] == "wb_result_output")

    source_download = marketplace_app_harness.client.get(
        f"/api/run-files/{source_file['id']}/download",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert source_download.status_code == 200
    assert source_download.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    result_download = marketplace_app_harness.client.get(
        f"/api/run-files/{result_file['id']}/download",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert result_download.status_code == 200
    workbook = load_workbook(BytesIO(result_download.content), data_only=False)
    sheet = workbook.worksheets[0]
    assert sheet["C2"].value == 15
    assert sheet["C3"].value == 40

    second_process_run = marketplace_app_harness.client.post(
        "/api/runs/process",
        headers=marketplace_app_harness.headers(manager_token),
        json={"store_id": store_id},
    )
    assert second_process_run.status_code == 200
    assert marketplace_app_harness.internal_drain_runs() == 2
    second_run_id = second_process_run.json()["id"]

    unavailable_page = marketplace_app_harness.client.get(
        f"/api/runs/{process_run_id}",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert unavailable_page.status_code == 200
    assert unavailable_page.json()["run"]["result_file_is_available"] is False
    assert unavailable_page.json()["run"]["result_file_unavailable_reason"] == "superseded"

    unavailable_download = marketplace_app_harness.client.get(
        f"/api/run-files/{result_file['id']}/download",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert unavailable_download.status_code == 410

    current_page = marketplace_app_harness.client.get(
        f"/api/runs/{second_run_id}",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert current_page.status_code == 200
    current_page_json = current_page.json()
    assert current_page_json["run"]["result_file_is_available"] is True
    current_result_file = next(item for item in current_page_json["files"] if item["file_role"] == "wb_result_output")

    current_download = marketplace_app_harness.client.get(
        f"/api/run-files/{current_result_file['id']}/download",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert current_download.status_code == 200

    history = marketplace_app_harness.client.get(
        "/api/history?page=1&page_size=25&store_id=%s&operation_type=process" % store_id,
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert history.status_code == 200
    history_ids = {item["run_id"] for item in history.json()["items"]}
    assert {process_run_id, second_run_id}.issubset(history_ids)

    logs = marketplace_app_harness.client.get(
        "/api/logs?page=1&page_size=25&event_type=old_result_removed_on_new_success",
        headers=marketplace_app_harness.headers(admin_token),
    )
    assert logs.status_code == 200
    supersede_logs = logs.json()["items"]
    assert len(supersede_logs) == 1
    assert supersede_logs[0]["run_id"] == process_run_id
    assert supersede_logs[0]["store_id"] == store_id
    assert supersede_logs[0]["module_code"] == "wb"
    assert supersede_logs[0]["payload_json"]["replacement_run_id"] == str(second_run_id)
    assert supersede_logs[0]["payload_json"]["reason"] == "superseded"

    all_logs = marketplace_app_harness.client.get(
        "/api/logs?page=1&page_size=100",
        headers=marketplace_app_harness.headers(admin_token),
    )
    assert all_logs.status_code == 200
    all_event_types = {item["event_type"] for item in all_logs.json()["items"]}
    assert "file_uploaded" in all_event_types
    assert "check_started" in all_event_types
    assert "check_finished" in all_event_types
    assert "process_started" in all_event_types
    assert "process_finished" in all_event_types
    assert "source_file_downloaded" in all_event_types
    assert "result_downloaded" in all_event_types


def test_ozon_process_end_to_end_acceptance(
    marketplace_app_harness,
    ozon_workbook_bytes: bytes,
) -> None:
    admin_token = marketplace_app_harness.login("admin", "admin-pass")
    manager_token = marketplace_app_harness.login("manager", "manager-pass")

    store = marketplace_app_harness.create_store(
        admin_token,
        name="AcceptanceOzonProcess",
        marketplace="ozon",
    )
    store_id = store["id"]
    marketplace_app_harness.grant_store_access(admin_token, user_id=2, store_id=store_id)

    uploaded = marketplace_app_harness.upload_temp_file(
        manager_token,
        store_id=store_id,
        module_code=ModuleCode.OZON,
        original_filename="ozon.xlsx",
        content=ozon_workbook_bytes,
    )
    assert uploaded.status_code == 200

    process_run = marketplace_app_harness.client.post(
        "/api/runs/process",
        headers=marketplace_app_harness.headers(manager_token),
        json={"store_id": store_id},
    )
    assert process_run.status_code == 200
    assert marketplace_app_harness.internal_drain_runs() == 2

    process_run_id = process_run.json()["id"]
    run_page = marketplace_app_harness.client.get(
        f"/api/runs/{process_run_id}",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert run_page.status_code == 200
    run_page_json = run_page.json()
    assert run_page_json["run"]["business_result"] == "completed"
    assert run_page_json["run"]["result_file_is_available"] is True

    result_file = next(item for item in run_page_json["files"] if item["file_role"] == "ozon_result_output")
    result_download = marketplace_app_harness.client.get(
        f"/api/run-files/{result_file['id']}/download",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert result_download.status_code == 200

    workbook = load_workbook(BytesIO(result_download.content), data_only=False)
    sheet = workbook["Товары и цены"]
    assert sheet["K4"].value == "Да"
    assert sheet["L4"].value == 700


def test_run_file_download_accepts_ui_cookie_session(
    marketplace_app_harness,
    wb_price_bytes: bytes,
    wb_promo_bytes: bytes,
) -> None:
    admin_token = marketplace_app_harness.login("admin", "admin-pass")
    manager_token = marketplace_app_harness.login("manager", "manager-pass")

    store = marketplace_app_harness.create_store(
        admin_token,
        name="CookieDownloadWB",
        marketplace="wb",
        wb_threshold_percent=60,
        wb_fallback_no_promo_percent=40,
        wb_fallback_over_threshold_percent=25,
    )
    store_id = int(store["id"])
    marketplace_app_harness.grant_store_access(admin_token, user_id=2, store_id=store_id)

    for filename, wb_file_kind, content in (
        ("price.xlsx", "price", wb_price_bytes),
        ("promo.xlsx", "promo", wb_promo_bytes),
    ):
        uploaded = marketplace_app_harness.upload_temp_file(
            manager_token,
            store_id=store_id,
            module_code=ModuleCode.WB,
            original_filename=filename,
            content=content,
            wb_file_kind=wb_file_kind,
        )
        assert uploaded.status_code == 200

    process_run = marketplace_app_harness.client.post(
        "/api/runs/process",
        headers=marketplace_app_harness.headers(manager_token),
        json={"store_id": store_id},
    )
    assert process_run.status_code == 200
    assert marketplace_app_harness.internal_drain_runs() == 2

    run_page = marketplace_app_harness.client.get(
        f"/api/runs/{process_run.json()['id']}",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert run_page.status_code == 200
    result_file = next(item for item in run_page.json()["files"] if item["file_role"] == "wb_result_output")

    marketplace_app_harness.use_ui_session(manager_token)
    cookie_download = marketplace_app_harness.client.get(
        f"/api/run-files/{result_file['id']}/download",
    )
    assert cookie_download.status_code == 200
    assert cookie_download.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    anonymous_client = marketplace_app_harness.client.__class__(marketplace_app_harness.client.app)
    missing_auth_download = anonymous_client.get(f"/api/run-files/{result_file['id']}/download")
    assert missing_auth_download.status_code == 401


def test_read_side_query_params_work_through_http_surface(
    marketplace_app_harness,
    wb_price_bytes: bytes,
    wb_promo_bytes: bytes,
) -> None:
    admin_token = marketplace_app_harness.login("admin", "admin-pass")
    manager_token = marketplace_app_harness.login("manager", "manager-pass")

    store = marketplace_app_harness.create_store(
        admin_token,
        name="ReadSideWB",
        marketplace="wb",
        wb_threshold_percent=60,
        wb_fallback_no_promo_percent=40,
        wb_fallback_over_threshold_percent=25,
    )
    marketplace_app_harness.grant_store_access(admin_token, user_id=2, store_id=store["id"])

    assert marketplace_app_harness.upload_temp_file(
        manager_token,
        store_id=store["id"],
        module_code=ModuleCode.WB,
        original_filename="price.xlsx",
        content=wb_price_bytes,
        wb_file_kind="price",
    ).status_code == 200
    assert marketplace_app_harness.upload_temp_file(
        manager_token,
        store_id=store["id"],
        module_code=ModuleCode.WB,
        original_filename="promo.xlsx",
        content=wb_promo_bytes,
        wb_file_kind="promo",
    ).status_code == 200

    check_run = marketplace_app_harness.client.post(
        "/api/runs/check",
        headers=marketplace_app_harness.headers(manager_token),
        json={"store_id": store["id"]},
    )
    assert check_run.status_code == 200
    assert marketplace_app_harness.internal_drain_runs() == 1

    history = marketplace_app_harness.client.get(
        f"/api/history?page=1&page_size=25&search=ReadSideWB&operation_type=check&store_id={store['id']}",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert history.status_code == 200
    assert history.json()["total_items"] == 1
    assert history.json()["items"][0]["public_run_number"] == check_run.json()["public_run_number"]

    logs = marketplace_app_harness.client.get(
        "/api/logs?page=1&page_size=25&event_type=check_finished&search=%s" % check_run.json()["public_run_number"],
        headers=marketplace_app_harness.headers(admin_token),
    )
    assert logs.status_code == 200
    assert any(item["public_run_number"] == check_run.json()["public_run_number"] for item in logs.json()["items"])

    detail = marketplace_app_harness.client.get(
        f"/api/audit/runs/{check_run.json()['id']}/detail?page=1&page_size=25&severity=info&sort_field=row_number&descending=false",
        headers=marketplace_app_harness.headers(manager_token),
    )
    assert detail.status_code == 200
    assert detail.json()["total_items"] >= 1
